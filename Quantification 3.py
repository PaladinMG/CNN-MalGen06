from __future__ import annotations

import argparse
from collections import defaultdict
from contextlib import ExitStack, contextmanager
import ctypes
from dataclasses import dataclass
from datetime import datetime
import gc
import importlib.metadata
import json
import logging
import math
import os
from pathlib import Path
import platform
import re
import shutil
import sys
import time
from typing import Iterator, Sequence

import numpy as np
from PIL import Image, UnidentifiedImageError
from scipy import ndimage as ndi
from scipy.spatial import cKDTree
from skimage import color, feature, filters, measure, morphology

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError as error:
    raise RuntimeError(
        "Quantification 3 requires openpyxl. Install the project requirements "
        "with 'python -m pip install -r requirements.txt'."
    ) from error


Image.MAX_IMAGE_PIXELS = None

CLASS_NAMES = (
    "Bone",
    "Fibrocartilage",
    "Cartilage",
    "Muscle",
    "Marrow",
    "Background",
)
CLASS_COLORS = {
    "Bone": "003B73",
    "Fibrocartilage": "79C7FF",
    "Cartilage": "8B0000",
    "Muscle": "FFB6C1",
    "Marrow": "800080",
    "Background": "FFFFFF",
}
RASTER_SUFFIXES = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}
ORIGINAL_SUFFIXES = RASTER_SUFFIXES | {".czi"}
SKIPPED_FILENAME_ENDINGS = ("20261.png", "20261.czi")
EXCEL_MAX_ROWS = 1_048_576
EPSILON = np.finfo(np.float32).eps
SCENE_PATTERN = re.compile(r"^(?P<stem>.+)__scene_(?P<scene>\d+)(?:__\d+)?$")
LOGGER = logging.getLogger("quantification3")
_LOGGING_CONFIGURED = False
TEXTURE_METRIC_FIELDS = (
    "Analysis pixels",
    "Red mean (0-1)",
    "Green mean (0-1)",
    "Blue mean (0-1)",
    "Red SD (0-1)",
    "Green SD (0-1)",
    "Blue SD (0-1)",
    "Combined RGB variation",
    "Red optical density mean",
    "Green optical density mean",
    "Blue optical density mean",
    "Hematoxylin stain mean",
    "Eosin stain mean",
    "Residual DAB stain mean",
    "Grayscale mean (0-1)",
    "Grayscale SD (0-1)",
    "Intensity Shannon entropy (bits)",
    "Gradient magnitude mean",
    "Gradient magnitude SD",
    "Local entropy mean (bits)",
    "Local entropy SD (bits)",
    "Structure-tensor coherence mean",
    "Structure-tensor orientation (degrees)",
    "Striation-orientation regularity",
    "Hessian ridge strength mean",
    "Hessian ridge strength P90",
    "GLCM contrast",
    "GLCM homogeneity",
    "GLCM energy",
    "GLCM correlation",
)


@dataclass(frozen=True)
class ScanInputs:
    name: str
    label_path: Path
    probability_path: Path | None
    grayscale_paths: tuple[Path | None, ...]
    boundary_path: Path | None
    original_path: Path | None
    original_scene: int | None


@dataclass
class AnalysisResults:
    scan_rows: list[dict[str, object]]
    class_rows: list[dict[str, object]]
    region_rows: list[dict[str, object]]
    pore_rows: list[dict[str, object]]
    interface_rows: list[dict[str, object]]
    spatial_rows: list[dict[str, object]]
    distance_band_rows: list[dict[str, object]]
    texture_rows: list[dict[str, object]]
    qc_rows: list[dict[str, object]]
    region_rows_omitted: int = 0
    pore_rows_omitted: int = 0


@dataclass
class NativeProbabilityChunkStats:
    expected_pixels: np.ndarray
    hard_confidence_sum: np.ndarray
    hard_confidence_min: np.ndarray
    hard_confidence_count: np.ndarray
    low_confidence_count: np.ndarray
    entropy_sum: float
    high_entropy_count: int
    low_margin_count: int
    bone_fibro_ambiguity_count: int


class ConsoleStatusLine:
    def __init__(self) -> None:
        self.stream = sys.stdout
        self.enabled = bool(getattr(self.stream, "isatty", lambda: False)())
        self.visible_length = 0

    def update(self, message: str) -> None:
        if not self.enabled:
            return
        padding = " " * max(0, self.visible_length - len(message))
        self.stream.write(f"\r{message}{padding}")
        self.stream.flush()
        self.visible_length = len(message)

    def clear(self) -> None:
        if not self.enabled or self.visible_length == 0:
            return
        self.stream.write("\r" + " " * self.visible_length + "\r")
        self.stream.flush()
        self.visible_length = 0


STATUS_LINE = ConsoleStatusLine()


class StatusAwareStreamHandler(logging.StreamHandler):
    def emit(self, record: logging.LogRecord) -> None:
        STATUS_LINE.clear()
        super().emit(record)


def log(message: str, level: int = logging.INFO) -> None:
    if _LOGGING_CONFIGURED:
        LOGGER.log(level, message)
    else:
        print(f"[Quantification 3] {message}", flush=True)


def update_status(message: str) -> None:
    status_message = f"[Quantification 3] {message}"
    if STATUS_LINE.enabled:
        if _LOGGING_CONFIGURED:
            log(f"Status | {message}", logging.DEBUG)
        STATUS_LINE.update(status_message)
    elif _LOGGING_CONFIGURED:
        log(message)
    else:
        print(status_message, flush=True)


def configure_logging(args: argparse.Namespace) -> Path:
    global _LOGGING_CONFIGURED
    args.output.parent.mkdir(parents=True, exist_ok=True)
    if args.log_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = args.output.parent / (
            f"quantification3_{timestamp}_pid{os.getpid()}.log"
        )
    else:
        log_path = args.log_file
    log_path.parent.mkdir(parents=True, exist_ok=True)

    for handler in list(LOGGER.handlers):
        handler.close()
        LOGGER.removeHandler(handler)
    LOGGER.setLevel(logging.DEBUG)
    LOGGER.propagate = False

    console_handler = StatusAwareStreamHandler(sys.stdout)
    console_handler.setLevel(getattr(logging, args.console_log_level))
    console_handler.setFormatter(
        logging.Formatter(
            "%(asctime)s | %(levelname)-8s | [Quantification 3] %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(
        logging.Formatter(
            "%(asctime)s.%(msecs)03d | %(levelname)-8s | "
            "pid=%(process)d | %(funcName)s:%(lineno)d | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    LOGGER.addHandler(console_handler)
    LOGGER.addHandler(file_handler)
    _LOGGING_CONFIGURED = True
    args.log_file = log_path
    log(f"Detailed log file: {log_path.resolve()}")
    return log_path


def package_version(distribution: str) -> str:
    try:
        return importlib.metadata.version(distribution)
    except importlib.metadata.PackageNotFoundError:
        return "not installed"


def cupy_package_version() -> str:
    accelerator = globals().get("ACCELERATOR")
    loaded_cupy = getattr(accelerator, "cupy", None)
    if loaded_cupy is not None:
        return str(getattr(loaded_cupy, "__version__", "unknown"))
    for distribution in ("cupy-cuda12x", "cupy-cuda13x", "cupy"):
        version = package_version(distribution)
        if version != "not installed":
            return version
    return "not installed"


def format_gib(value: int | None) -> str:
    return "unavailable" if value is None else f"{value / 1024 ** 3:.3f} GiB"


def file_size_or_unavailable(path: Path | None) -> int | str:
    if path is None:
        return "unavailable"
    try:
        return path.stat().st_size
    except OSError:
        return "unavailable"


class AccelerationBackend:
    """Optional CUDA acceleration with transparent, logged CPU fallbacks."""

    def __init__(self) -> None:
        self.enabled = False
        self.requested_device = "cpu"
        self.resolved_device = "cpu"
        self.device_index = 0
        self.device_name = "CPU"
        self.memory_fraction = 0.80
        self.configured_chunk_rows = 0
        self.minimum_pixels = 262_144
        self.torch = None
        self.torch_device = None
        self.cupy = None
        self.cupy_ndi = None
        self.disabled_stages: set[str] = set()
        self.reported_stages: set[str] = set()
        self.fallback_count = 0

    @property
    def torch_cuda_available(self) -> bool:
        return self.torch is not None and self.torch_device is not None

    @property
    def cupy_available(self) -> bool:
        return self.cupy is not None and self.cupy_ndi is not None

    def configure(self, args: argparse.Namespace) -> None:
        self.requested_device = args.device
        self.device_index = args.cuda_device
        self.memory_fraction = args.gpu_memory_fraction
        self.configured_chunk_rows = args.gpu_chunk_rows
        self.minimum_pixels = args.gpu_min_pixels
        if args.device == "cpu":
            self._record_args(args)
            log("Compute backend | requested=cpu | resolved=cpu")
            return

        torch_error: Exception | None = None
        try:
            import torch

            self.torch = torch
            if torch.cuda.is_available():
                device_count = torch.cuda.device_count()
                if not 0 <= args.cuda_device < device_count:
                    raise RuntimeError(
                        f"--cuda-device {args.cuda_device} is invalid; "
                        f"PyTorch reports {device_count} CUDA device(s)"
                    )
                torch.cuda.set_device(args.cuda_device)
                self.torch_device = torch.device(
                    "cuda", index=args.cuda_device
                )
                self.device_name = torch.cuda.get_device_name(args.cuda_device)
        except Exception as error:
            torch_error = error
            self.torch = None
            self.torch_device = None

        cupy_error: Exception | None = None
        if not args.disable_cupy:
            try:
                import cupy
                from cupyx.scipy import ndimage as cupy_ndi

                device_count = int(cupy.cuda.runtime.getDeviceCount())
                if not 0 <= args.cuda_device < device_count:
                    raise RuntimeError(
                        f"--cuda-device {args.cuda_device} is invalid; "
                        f"CuPy reports {device_count} CUDA device(s)"
                    )
                cupy.cuda.Device(args.cuda_device).use()
                test_value = cupy.asarray([1.0], dtype=cupy.float32)
                cupy.cuda.get_current_stream().synchronize()
                del test_value
                self.cupy = cupy
                self.cupy_ndi = cupy_ndi
                try:
                    cupy.get_default_memory_pool().set_limit(
                        fraction=args.gpu_memory_fraction
                    )
                except (AttributeError, TypeError):
                    log(
                        "CuPy memory-pool fraction limit is unavailable in "
                        "this CuPy version",
                        logging.DEBUG,
                    )
                if self.device_name == "CPU":
                    properties = cupy.cuda.runtime.getDeviceProperties(
                        args.cuda_device
                    )
                    raw_name = properties.get("name", "CUDA GPU")
                    self.device_name = (
                        raw_name.decode(errors="replace")
                        if isinstance(raw_name, bytes) else str(raw_name)
                    )
            except Exception as error:
                cupy_error = error
                self.cupy = None
                self.cupy_ndi = None

        self.enabled = self.torch_cuda_available or self.cupy_available
        self.resolved_device = (
            f"cuda:{args.cuda_device}" if self.enabled else "cpu"
        )
        if args.device == "cuda" and not self.enabled:
            details = []
            if torch_error is not None:
                details.append(f"PyTorch: {torch_error}")
            else:
                details.append("PyTorch was built without usable CUDA support")
            if args.disable_cupy:
                details.append("CuPy disabled by --disable-cupy")
            elif cupy_error is not None:
                details.append(f"CuPy: {cupy_error}")
            raise RuntimeError(
                "CUDA was requested but no CUDA backend could be initialized. "
                + " | ".join(details)
            )

        if not self.enabled:
            log(
                "Compute backend | "
                f"requested={args.device} | resolved=cpu | device=CPU"
            )
            log(
                "CUDA is unavailable; using the CPU backend. Add --device cuda "
                "to require CUDA instead of allowing this fallback.",
                logging.WARNING,
            )
        else:
            torch_status = (
                f"enabled ({package_version('torch')}; runtime CUDA "
                f"{getattr(self.torch.version, 'cuda', 'unknown')})"
                if self.torch_cuda_available else "unavailable"
            )
            cupy_status = (
                f"enabled ({cupy_package_version()})"
                if self.cupy_available else "unavailable"
            )
            log(
                "Compute backend | "
                f"requested={args.device} | resolved={self.resolved_device} | "
                f"device={self.device_name} | PyTorch CUDA={torch_status} | "
                f"CuPy ndimage={cupy_status} | "
                f"memory_fraction={self.memory_fraction:.2f} | "
                f"gpu_min_pixels={self.minimum_pixels:,}"
            )
            if not self.cupy_available and not args.disable_cupy:
                log(
                    "CuPy is unavailable: probability reductions can still use "
                    "PyTorch CUDA, but distance transforms, morphology, and "
                    "texture filters will remain on the CPU. Install "
                    "requirements-gpu.txt for the full CUDA path.",
                    logging.WARNING,
                )
                if cupy_error is not None:
                    log(f"CuPy initialization detail: {cupy_error}", logging.DEBUG)
        self._record_args(args)

    def _record_args(self, args: argparse.Namespace) -> None:
        args.resolved_device = self.resolved_device
        args.cuda_device_name = self.device_name
        args.pytorch_cuda_enabled = self.torch_cuda_available
        args.cupy_enabled = self.cupy_available

    def should_use_torch(self, pixel_count: int, stage: str) -> bool:
        return (
            self.torch_cuda_available
            and pixel_count >= self.minimum_pixels
            and stage not in self.disabled_stages
        )

    def should_use_cupy(self, pixel_count: int, stage: str) -> bool:
        return (
            self.cupy_available
            and pixel_count >= self.minimum_pixels
            and stage not in self.disabled_stages
        )

    def available_gpu_bytes(self) -> int | None:
        try:
            if self.cupy_available:
                free_bytes, _ = self.cupy.cuda.runtime.memGetInfo()
                return int(free_bytes)
            if self.torch_cuda_available:
                free_bytes, _ = self.torch.cuda.mem_get_info(self.device_index)
                return int(free_bytes)
        except Exception:
            return None
        return None

    def choose_chunk_rows(
        self,
        width: int,
        height: int,
        bytes_per_pixel: int,
        halo_rows: int = 0,
    ) -> int:
        if self.configured_chunk_rows > 0:
            return min(height, self.configured_chunk_rows)
        free_bytes = self.available_gpu_bytes()
        if free_bytes is None:
            return min(height, 512)
        budget = int(free_bytes * min(self.memory_fraction, 0.80) * 0.65)
        rows = budget // max(1, width * bytes_per_pixel) - 2 * halo_rows
        return max(1, min(height, 4096, int(rows)))

    def disable_stage(self, stage: str, error: Exception) -> None:
        self.disabled_stages.add(stage)
        self.fallback_count += 1
        log(
            f"GPU stage fallback | stage={stage} | error={type(error).__name__}: "
            f"{error} | subsequent work for this stage will use CPU",
            logging.WARNING,
        )
        LOGGER.debug("GPU fallback traceback", exc_info=True)
        self.release_memory()

    def release_memory(self) -> None:
        if self.torch_cuda_available:
            try:
                self.torch.cuda.empty_cache()
            except Exception:
                pass
        if self.cupy_available:
            try:
                self.cupy.get_default_memory_pool().free_all_blocks()
                self.cupy.get_default_pinned_memory_pool().free_all_blocks()
            except Exception:
                pass

    def resource_text(self) -> str:
        if not self.enabled:
            return "gpu=disabled"
        parts = [f"gpu={self.resolved_device}"]
        try:
            if self.torch_cuda_available:
                torch = self.torch
                free_bytes, total_bytes = torch.cuda.mem_get_info(
                    self.device_index
                )
                parts.extend(
                    [
                        f"gpu_free={format_gib(int(free_bytes))}",
                        f"gpu_total={format_gib(int(total_bytes))}",
                        f"torch_allocated={format_gib(torch.cuda.memory_allocated(self.device_index))}",
                        f"torch_reserved={format_gib(torch.cuda.memory_reserved(self.device_index))}",
                    ]
                )
            elif self.cupy_available:
                free_bytes, total_bytes = self.cupy.cuda.runtime.memGetInfo()
                parts.extend(
                    [
                        f"gpu_free={format_gib(int(free_bytes))}",
                        f"gpu_total={format_gib(int(total_bytes))}",
                    ]
                )
        except Exception as error:
            parts.append(f"gpu_memory=unavailable({error})")
        return " | ".join(parts)

    def ndimage(
        self,
        operation: str,
        array: np.ndarray,
        *args: object,
        **kwargs: object,
    ) -> object:
        cpu_function = getattr(ndi, operation)
        stage = f"cupy_{operation}"
        if not self.should_use_cupy(array.size, stage):
            return cpu_function(array, *args, **kwargs)
        cupy = self.cupy

        def to_gpu(value: object) -> object:
            return cupy.asarray(value) if isinstance(value, np.ndarray) else value

        try:
            gpu_args = tuple(to_gpu(value) for value in args)
            gpu_kwargs = {
                key: to_gpu(value) for key, value in kwargs.items()
            }
            gpu_array = cupy.asarray(array)
            gpu_function = getattr(self.cupy_ndi, operation)
            if operation == "distance_transform_edt":
                gpu_kwargs.setdefault("float64_distances", True)
            try:
                result = gpu_function(gpu_array, *gpu_args, **gpu_kwargs)
            except TypeError:
                if operation != "distance_transform_edt":
                    raise
                gpu_kwargs.pop("float64_distances", None)
                result = gpu_function(gpu_array, *gpu_args, **gpu_kwargs)

            def to_cpu(value: object) -> object:
                if isinstance(value, cupy.ndarray):
                    return cupy.asnumpy(value)
                if isinstance(value, cupy.generic):
                    return value.item()
                return value

            if isinstance(result, tuple):
                cpu_result = tuple(to_cpu(value) for value in result)
            else:
                cpu_result = to_cpu(result)
            if stage not in self.reported_stages:
                self.reported_stages.add(stage)
                log(
                    f"GPU stage active | stage={stage} | "
                    f"array_shape={array.shape} | array_dtype={array.dtype}",
                    logging.DEBUG,
                )
            return cpu_result
        except Exception as error:
            self.disable_stage(stage, error)
            return cpu_function(array, *args, **kwargs)


ACCELERATOR = AccelerationBackend()


def accelerated_distance_transform_edt(
    array: np.ndarray,
    *args: object,
    **kwargs: object,
) -> np.ndarray:
    return ACCELERATOR.ndimage(
        "distance_transform_edt", array, *args, **kwargs
    )


def accelerated_binary_erosion(
    array: np.ndarray,
    *args: object,
    **kwargs: object,
) -> np.ndarray:
    return ACCELERATOR.ndimage("binary_erosion", array, *args, **kwargs)


def accelerated_binary_dilation(
    array: np.ndarray,
    *args: object,
    **kwargs: object,
) -> np.ndarray:
    return ACCELERATOR.ndimage("binary_dilation", array, *args, **kwargs)


def accelerated_binary_fill_holes(
    array: np.ndarray,
    *args: object,
    **kwargs: object,
) -> np.ndarray:
    return ACCELERATOR.ndimage("binary_fill_holes", array, *args, **kwargs)


def accelerated_label(
    array: np.ndarray,
    *args: object,
    **kwargs: object,
) -> tuple[np.ndarray, int]:
    labels, count = ACCELERATOR.ndimage("label", array, *args, **kwargs)
    return labels, int(count)


def accelerated_convolve(
    array: np.ndarray,
    *args: object,
    **kwargs: object,
) -> np.ndarray:
    return ACCELERATOR.ndimage("convolve", array, *args, **kwargs)


def available_memory_bytes() -> int | None:
    if os.name == "nt":
        class MemoryStatusEx(ctypes.Structure):
            _fields_ = [
                ("dwLength", ctypes.c_ulong),
                ("dwMemoryLoad", ctypes.c_ulong),
                ("ullTotalPhys", ctypes.c_ulonglong),
                ("ullAvailPhys", ctypes.c_ulonglong),
                ("ullTotalPageFile", ctypes.c_ulonglong),
                ("ullAvailPageFile", ctypes.c_ulonglong),
                ("ullTotalVirtual", ctypes.c_ulonglong),
                ("ullAvailVirtual", ctypes.c_ulonglong),
                ("ullAvailExtendedVirtual", ctypes.c_ulonglong),
            ]

        status = MemoryStatusEx()
        status.dwLength = ctypes.sizeof(status)
        if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(status)):
            return int(status.ullAvailPhys)
        return None
    if hasattr(os, "sysconf"):
        try:
            return int(
                os.sysconf("SC_AVPHYS_PAGES") * os.sysconf("SC_PAGE_SIZE")
            )
        except (ValueError, OSError):
            return None
    return None


def process_memory_bytes() -> tuple[int | None, int | None]:
    if os.name == "nt":
        class ProcessMemoryCounters(ctypes.Structure):
            _fields_ = [
                ("cb", ctypes.c_ulong),
                ("PageFaultCount", ctypes.c_ulong),
                ("PeakWorkingSetSize", ctypes.c_size_t),
                ("WorkingSetSize", ctypes.c_size_t),
                ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                ("QuotaPagedPoolUsage", ctypes.c_size_t),
                ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                ("PagefileUsage", ctypes.c_size_t),
                ("PeakPagefileUsage", ctypes.c_size_t),
            ]

        counters = ProcessMemoryCounters()
        counters.cb = ctypes.sizeof(counters)
        process = ctypes.windll.kernel32.GetCurrentProcess()
        if ctypes.windll.psapi.GetProcessMemoryInfo(
            process, ctypes.byref(counters), counters.cb
        ):
            return int(counters.WorkingSetSize), int(counters.PeakWorkingSetSize)
    return None, None


def log_resource_snapshot(stage: str, output_path: Path) -> None:
    current_memory, peak_memory = process_memory_bytes()
    available_memory = available_memory_bytes()
    try:
        disk = shutil.disk_usage(output_path.parent)
        disk_free = disk.free
    except OSError:
        disk_free = None
    log(
        "Resource snapshot | "
        f"stage={stage} | rss={format_gib(current_memory)} | "
        f"peak_rss={format_gib(peak_memory)} | "
        f"available_ram={format_gib(available_memory)} | "
        f"output_disk_free={format_gib(disk_free)} | "
        f"{ACCELERATOR.resource_text()}",
        logging.DEBUG,
    )


def log_startup(args: argparse.Namespace) -> None:
    log(
        "Runtime | "
        f"python={platform.python_version()} | platform={platform.platform()} | "
        f"pid={os.getpid()} | cwd={Path.cwd()}"
    )
    log(
        "Libraries | "
        f"numpy={package_version('numpy')} | Pillow={package_version('Pillow')} | "
        f"scipy={package_version('scipy')} | "
        f"scikit-image={package_version('scikit-image')} | "
        f"openpyxl={package_version('openpyxl')} | "
        f"torch={package_version('torch')} | cupy={cupy_package_version()}"
    )
    log(f"Command | {' '.join(sys.argv)}", logging.DEBUG)
    for name, value in sorted(vars(args).items()):
        log(f"Argument | {name}={value}", logging.DEBUG)
    log_resource_snapshot("startup", args.output)


def finite_or_none(value: object) -> object:
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def json_measurement_value(value: object) -> object:
    value = finite_or_none(value)
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def log_measurement_record(
    category: str,
    record_index: int,
    record: dict[str, object],
) -> None:
    payload = {
        key: json_measurement_value(value)
        for key, value in record.items()
    }
    log(
        f"Measurement | category={category} | record={record_index} | "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        logging.DEBUG,
    )


def log_all_measurements(
    results: AnalysisResults,
    mode: str,
) -> None:
    if mode == "none":
        log("Per-measurement logging disabled by --measurement-log-mode none")
        return
    groups: list[tuple[str, list[dict[str, object]]]] = [
        ("scan_summary", results.scan_rows),
        ("class_summary", results.class_rows),
        ("interfaces", results.interface_rows),
        ("spatial_distances", results.spatial_rows),
        ("distance_bands", results.distance_band_rows),
        ("texture_color", results.texture_rows),
        ("quality_control", results.qc_rows),
    ]
    if mode == "all":
        groups[2:2] = [
            ("region_details", results.region_rows),
            ("pore_details", results.pore_rows),
        ]
    record_total = sum(len(records) for _, records in groups)
    log(
        f"Writing {record_total:,} structured measurement record(s) "
        f"to the detailed log (mode={mode})"
    )
    for category, records in groups:
        log(
            f"Measurement category | category={category} | "
            f"records={len(records):,}",
            logging.DEBUG,
        )
        for record_index, record in enumerate(records, start=1):
            log_measurement_record(category, record_index, record)
    log("Structured measurement logging complete")


def safe_stat(values: Sequence[float] | np.ndarray, statistic: str) -> float | None:
    array = np.asarray(values, dtype=np.float64)
    array = array[np.isfinite(array)]
    if not array.size:
        return None
    if statistic == "mean":
        return float(array.mean())
    if statistic == "std":
        return float(array.std())
    if statistic == "min":
        return float(array.min())
    if statistic == "max":
        return float(array.max())
    if statistic == "median":
        return float(np.median(array))
    if statistic.startswith("p"):
        return float(np.percentile(array, float(statistic[1:])))
    raise ValueError(f"Unknown statistic: {statistic}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Measure tissue morphology, topology, interfaces, spatial "
            "organization, image texture, and prediction uncertainty."
        )
    )
    parser.add_argument("--prediction-dir", type=Path, default=Path("prediction"))
    parser.add_argument(
        "--output",
        type=Path,
        help="Output .xlsx path (default: <prediction-dir>/Quantification 3.xlsx).",
    )
    parser.add_argument(
        "--original-dir",
        type=Path,
        help=(
            "Optional directory containing matching RGB raster or CZI originals. "
            "Enables color, stain, gradient, and texture measurements."
        ),
    )
    parser.add_argument("--pixel-width-um", type=float, default=0.22)
    parser.add_argument("--pixel-height-um", type=float, default=0.22)
    parser.add_argument(
        "--analysis-downsample",
        type=int,
        default=1,
        help=(
            "Downsampling used only for morphology/texture-heavy metrics. "
            "Areas and probability totals remain native-resolution."
        ),
    )
    parser.add_argument("--chunk-rows", type=int, default=128)
    parser.add_argument("--low-confidence-threshold", type=float, default=0.50)
    parser.add_argument("--uncertainty-margin", type=float, default=0.15)
    parser.add_argument(
        "--high-entropy-threshold",
        type=float,
        default=0.75,
        help="Threshold for entropy normalized to the range 0-1.",
    )
    parser.add_argument("--proximity-um", type=float, default=10.0)
    parser.add_argument("--distance-bin-um", type=float, default=50.0)
    parser.add_argument("--texture-levels", type=int, default=32)
    parser.add_argument("--texture-entropy-radius", type=int, default=5)
    parser.add_argument(
        "--device",
        choices=("auto", "cpu", "cuda"),
        default="auto",
        help=(
            "Compute device. 'auto' uses CUDA when available, 'cpu' disables "
            "GPU work, and 'cuda' fails if no CUDA backend can initialize."
        ),
    )
    parser.add_argument(
        "--cuda-device",
        type=int,
        default=0,
        help="Zero-based CUDA device index.",
    )
    parser.add_argument(
        "--gpu-memory-fraction",
        type=float,
        default=0.80,
        help=(
            "Fraction of currently available VRAM targeted by automatic GPU "
            "chunking and the CuPy memory pool."
        ),
    )
    parser.add_argument(
        "--gpu-chunk-rows",
        type=int,
        default=0,
        help=(
            "Rows per GPU chunk. Zero selects a VRAM-aware value automatically."
        ),
    )
    parser.add_argument(
        "--gpu-min-pixels",
        type=int,
        default=262_144,
        help=(
            "Minimum array pixels before CUDA is used; avoids transfer overhead "
            "for small operations."
        ),
    )
    parser.add_argument(
        "--disable-cupy",
        action="store_true",
        help=(
            "Use only the PyTorch CUDA path; distance transforms, morphology, "
            "and texture filtering then remain on CPU."
        ),
    )
    parser.add_argument(
        "--max-region-rows",
        type=int,
        default=500_000,
        help="Maximum connected-region detail rows retained in the workbook.",
    )
    parser.add_argument(
        "--max-pore-rows",
        type=int,
        default=500_000,
        help="Maximum pore detail rows retained in the workbook.",
    )
    parser.add_argument(
        "--skip-curvature",
        action="store_true",
        help="Skip signed-distance curvature calculations to reduce memory/time.",
    )
    parser.add_argument(
        "--skip-skeleton",
        action="store_true",
        help="Skip skeleton, branch, junction, and local-thickness calculations.",
    )
    parser.add_argument(
        "--log-file",
        type=Path,
        help=(
            "Detailed UTF-8 log path. By default, a timestamped log is created "
            "beside the output workbook."
        ),
    )
    parser.add_argument(
        "--console-log-level",
        choices=("INFO", "DEBUG"),
        default="INFO",
        help="Console verbosity; the file log always includes DEBUG records.",
    )
    parser.add_argument(
        "--measurement-log-mode",
        choices=("all", "summary", "none"),
        default="all",
        help=(
            "Measurements written to the detailed log. 'all' includes every "
            "retained region and pore; 'summary' omits object-detail rows."
        ),
    )
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    if not args.prediction_dir.is_dir():
        parser.error(f"Prediction directory does not exist: {args.prediction_dir}")
    if args.original_dir is not None and not args.original_dir.is_dir():
        parser.error(f"Original-image directory does not exist: {args.original_dir}")
    if args.pixel_width_um <= 0 or args.pixel_height_um <= 0:
        parser.error("Pixel dimensions must be positive")
    if args.analysis_downsample < 1:
        parser.error("--analysis-downsample must be at least 1")
    if args.chunk_rows < 1:
        parser.error("--chunk-rows must be at least 1")
    if not 0 <= args.low_confidence_threshold <= 1:
        parser.error("--low-confidence-threshold must be between 0 and 1")
    if not 0 <= args.uncertainty_margin <= 1:
        parser.error("--uncertainty-margin must be between 0 and 1")
    if not 0 <= args.high_entropy_threshold <= 1:
        parser.error("--high-entropy-threshold must be between 0 and 1")
    if args.proximity_um < 0 or args.distance_bin_um <= 0:
        parser.error("Proximity must be nonnegative and distance bin size positive")
    if args.texture_levels < 4 or args.texture_levels > 256:
        parser.error("--texture-levels must be between 4 and 256")
    if args.texture_entropy_radius < 1:
        parser.error("--texture-entropy-radius must be at least 1")
    if args.cuda_device < 0:
        parser.error("--cuda-device cannot be negative")
    if not 0 < args.gpu_memory_fraction <= 1:
        parser.error("--gpu-memory-fraction must be greater than 0 and at most 1")
    if args.gpu_chunk_rows < 0:
        parser.error("--gpu-chunk-rows cannot be negative")
    if args.gpu_min_pixels < 1:
        parser.error("--gpu-min-pixels must be at least 1")
    if args.max_region_rows < 0 or args.max_pore_rows < 0:
        parser.error("Detail row limits cannot be negative")

    args.output = args.output or args.prediction_dir / "Quantification 3.xlsx"
    if args.output.suffix.lower() != ".xlsx":
        parser.error("--output must end in .xlsx")
    if args.output.exists() and not args.overwrite:
        parser.error(f"Output already exists; add --overwrite: {args.output}")
    return args


def find_optional_path(directory: Path, name: str, suffix: str) -> Path | None:
    path = directory / f"{name}{suffix}"
    return path if path.is_file() else None


def should_skip_file(path: Path) -> bool:
    return path.name.casefold().endswith(SKIPPED_FILENAME_ENDINGS)


def build_original_index(root: Path | None) -> dict[str, list[Path]]:
    if root is None:
        log("Original-image indexing skipped: --original-dir was not supplied", logging.DEBUG)
        return {}
    started = time.perf_counter()
    log(f"Indexing original images under: {root.resolve()}")
    index: dict[str, list[Path]] = defaultdict(list)
    candidate_count = 0
    skipped_count = 0
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in ORIGINAL_SUFFIXES:
            continue
        candidate_count += 1
        if should_skip_file(path):
            log(f"Skipping excluded original file: {path}")
            skipped_count += 1
            continue
        index[path.stem.casefold()].append(path)
    indexed_count = sum(len(paths) for paths in index.values())
    log(
        "Original-image index complete | "
        f"candidates={candidate_count:,} | indexed={indexed_count:,} | "
        f"skipped={skipped_count:,} | unique_stems={len(index):,} | "
        f"elapsed_seconds={time.perf_counter() - started:.3f}"
    )
    return index


def match_original(
    prediction_name: str,
    index: dict[str, list[Path]],
) -> tuple[Path | None, int | None, str | None]:
    if not index:
        return None, None, "No --original-dir was supplied"
    scene_match = SCENE_PATTERN.match(prediction_name)
    lookup_name = (
        scene_match.group("stem") if scene_match else prediction_name
    ).casefold()
    scene_index = int(scene_match.group("scene")) if scene_match else None
    candidates = index.get(prediction_name.casefold(), [])
    if not candidates:
        candidates = index.get(lookup_name, [])
    if not candidates:
        return None, scene_index, "No matching original image was found"
    if len(candidates) > 1:
        return (
            None,
            scene_index,
            "More than one matching original image was found: "
            + "; ".join(str(path) for path in candidates[:5]),
        )
    return candidates[0], scene_index, None


def discover_scans(prediction_dir: Path, original_dir: Path | None) -> list[ScanInputs]:
    started = time.perf_counter()
    full_dir = prediction_dir / "Full Segmentations"
    log(f"Discovering full segmentations under: {full_dir.resolve()}")
    if not full_dir.is_dir():
        raise FileNotFoundError(f"Missing Full Segmentations directory: {full_dir}")
    candidate_label_paths = sorted(
        path for path in full_dir.iterdir()
        if path.is_file() and path.suffix.lower() in RASTER_SUFFIXES
    )
    label_paths = []
    skipped_label_count = 0
    for path in candidate_label_paths:
        if should_skip_file(path):
            log(f"Skipping excluded segmentation file: {path}")
            skipped_label_count += 1
        else:
            label_paths.append(path)
    if not label_paths:
        raise FileNotFoundError(
            f"No eligible full segmentation images found in {full_dir}. "
            "Files ending in 20261.png are intentionally skipped."
        )

    original_index = build_original_index(original_dir)
    scans: list[ScanInputs] = []
    for label_path in label_paths:
        name = label_path.stem
        probability_path = find_optional_path(
            prediction_dir / "Probabilities", name, ".npy"
        )
        grayscale_paths = tuple(
            find_optional_path(
                prediction_dir / class_name / "Grayscale", name, ".png"
            )
            for class_name in CLASS_NAMES
        )
        boundary_path = find_optional_path(
            prediction_dir / "Boundary", name, ".png"
        )
        original_path, original_scene, original_warning = match_original(
            name, original_index
        )
        if original_warning and original_dir is not None:
            log(f"{name}: {original_warning}")
        if probability_path is None and any(path is None for path in grayscale_paths):
            missing = [
                class_name
                for class_name, path in zip(CLASS_NAMES, grayscale_paths)
                if path is None
            ]
            log(
                f"{name}: probability-derived metrics unavailable because no "
                f".npy exists and grayscale maps are missing for {', '.join(missing)}"
            )
        scans.append(
            ScanInputs(
                name=name,
                label_path=label_path,
                probability_path=probability_path,
                grayscale_paths=grayscale_paths,
                boundary_path=boundary_path,
                original_path=original_path,
                original_scene=original_scene,
            )
        )
        log(
            "Discovered scan inputs | "
            f"scan={name} | label={label_path} | "
            f"probability={probability_path or 'unavailable'} | "
            f"grayscale_maps={sum(path is not None for path in grayscale_paths)}/"
            f"{len(CLASS_NAMES)} | boundary={boundary_path or 'unavailable'} | "
            f"original={original_path or 'unavailable'} | scene={original_scene}",
            logging.DEBUG,
        )
    log(
        "Scan discovery complete | "
        f"candidates={len(candidate_label_paths):,} | scans={len(scans):,} | "
        f"skipped={skipped_label_count:,} | "
        f"elapsed_seconds={time.perf_counter() - started:.3f}"
    )
    return scans


class ProbabilitySource:
    def __init__(self, scan: ScanInputs, height: int, width: int) -> None:
        self.scan = scan
        self.height = height
        self.width = width
        self.memmap: np.ndarray | None = None
        self.images: list[Image.Image] = []
        self.kind: str | None = None

    def __enter__(self) -> "ProbabilitySource":
        if self.scan.probability_path is not None:
            log(
                f"{self.scan.name}: opening float32 probability volume | "
                f"path={self.scan.probability_path} | "
                f"bytes={file_size_or_unavailable(self.scan.probability_path)}",
                logging.DEBUG,
            )
            array = np.load(self.scan.probability_path, mmap_mode="r")
            expected_shape = (len(CLASS_NAMES), self.height, self.width)
            if array.shape != expected_shape:
                raise ValueError(
                    f"{self.scan.probability_path} has shape {array.shape}; "
                    f"expected {expected_shape}"
                )
            self.memmap = array
            self.kind = "float32 probability volume"
            log(
                f"{self.scan.name}: probability volume ready | "
                f"shape={array.shape} | dtype={array.dtype}",
                logging.DEBUG,
            )
            return self
        if all(path is not None for path in self.scan.grayscale_paths):
            log(
                f"{self.scan.name}: opening six grayscale probability maps",
                logging.DEBUG,
            )
            stack = ExitStack()
            try:
                self.images = [
                    stack.enter_context(Image.open(path))
                    for path in self.scan.grayscale_paths
                    if path is not None
                ]
                for image, path in zip(self.images, self.scan.grayscale_paths):
                    if image.size != (self.width, self.height):
                        raise ValueError(
                            f"{path} has size {image.size}; expected "
                            f"{(self.width, self.height)}"
                        )
                self._stack = stack
                self.kind = "8-bit grayscale probability maps"
                log(
                    f"{self.scan.name}: grayscale probability maps ready | "
                    f"dimensions={self.width}x{self.height}",
                    logging.DEBUG,
                )
                return self
            except Exception:
                stack.close()
                raise
        self.kind = None
        log(
            f"{self.scan.name}: no complete probability source is available",
            logging.WARNING,
        )
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        if hasattr(self, "_stack"):
            self._stack.close()
        self.images.clear()
        if self.memmap is not None:
            mmap = getattr(self.memmap, "_mmap", None)
            if mmap is not None:
                mmap.close()
            self.memmap = None
        log(
            f"{self.scan.name}: probability resources closed",
            logging.DEBUG,
        )

    @property
    def available(self) -> bool:
        return self.kind is not None

    def native_chunk(self, top: int, bottom: int) -> np.ndarray:
        if not self.available:
            raise RuntimeError("Probability data are unavailable")
        if self.memmap is not None:
            return np.asarray(self.memmap[:, top:bottom], dtype=np.float32)
        arrays = [
            np.asarray(
                image.crop((0, top, self.width, bottom)).convert("L"),
                dtype=np.float32,
            )
            for image in self.images
        ]
        return np.stack(arrays, axis=0) / np.float32(255.0)

    def analysis_class(
        self,
        class_id: int,
        analysis_height: int,
        analysis_width: int,
        output_chunk_rows: int,
    ) -> np.ndarray:
        if not self.available:
            raise RuntimeError("Probability data are unavailable")
        output = np.empty((analysis_height, analysis_width), dtype=np.float32)
        if self.memmap is not None:
            source_rows = np.minimum(
                (
                    (np.arange(analysis_height, dtype=np.float64) + 0.5)
                    * self.height
                    / analysis_height
                ).astype(np.int64),
                self.height - 1,
            )
            source_cols = np.minimum(
                (
                    (np.arange(analysis_width, dtype=np.float64) + 0.5)
                    * self.width
                    / analysis_width
                ).astype(np.int64),
                self.width - 1,
            )
            source = self.memmap[class_id]
            for out_top in range(0, analysis_height, output_chunk_rows):
                out_bottom = min(out_top + output_chunk_rows, analysis_height)
                rows = source_rows[out_top:out_bottom]
                output[out_top:out_bottom] = source[
                    rows[:, None], source_cols[None, :]
                ]
            return output
        resized = self.images[class_id].resize(
            (analysis_width, analysis_height), Image.Resampling.NEAREST
        ).convert("L")
        try:
            output[:] = np.asarray(resized, dtype=np.float32) / np.float32(255.0)
        finally:
            resized.close()
        return output

    def analysis_chunk(
        self,
        top: int,
        bottom: int,
        analysis_height: int,
        analysis_width: int,
    ) -> np.ndarray:
        arrays = []
        source_top = math.floor(top * self.height / analysis_height)
        source_bottom = math.ceil(bottom * self.height / analysis_height)
        source_bottom = max(source_top + 1, min(source_bottom, self.height))
        for class_id in range(len(CLASS_NAMES)):
            if self.memmap is not None:
                source_rows = np.minimum(
                    (
                        (np.arange(top, bottom, dtype=np.float64) + 0.5)
                        * self.height
                        / analysis_height
                    ).astype(np.int64),
                    self.height - 1,
                )
                source_cols = np.minimum(
                    (
                        (np.arange(analysis_width, dtype=np.float64) + 0.5)
                        * self.width
                        / analysis_width
                    ).astype(np.int64),
                    self.width - 1,
                )
                arrays.append(
                    np.asarray(
                        self.memmap[class_id][
                            source_rows[:, None], source_cols[None, :]
                        ],
                        dtype=np.float32,
                    )
                )
            else:
                crop = self.images[class_id].crop(
                    (0, source_top, self.width, source_bottom)
                )
                resized = crop.resize(
                    (analysis_width, bottom - top), Image.Resampling.NEAREST
                ).convert("L")
                try:
                    arrays.append(
                        np.asarray(resized, dtype=np.float32)
                        / np.float32(255.0)
                    )
                finally:
                    resized.close()
                    crop.close()
        return np.stack(arrays, axis=0)


def palette_class_remap(
    image: Image.Image,
    label_path: Path,
) -> np.ndarray | None:
    if image.mode != "P":
        return None
    raw_palette = image.getpalette()
    if raw_palette is None:
        raise ValueError(f"{label_path} is a palette image without a palette")
    palette = np.asarray(raw_palette, dtype=np.uint8).reshape((-1, 3))
    remap = np.full(len(palette), 255, dtype=np.uint8)
    for class_id, class_name in enumerate(CLASS_NAMES):
        color_hex = CLASS_COLORS[class_name]
        expected_color = np.asarray(
            [
                int(color_hex[0:2], 16),
                int(color_hex[2:4], 16),
                int(color_hex[4:6], 16),
            ],
            dtype=np.uint8,
        )
        matching_indices = np.flatnonzero(
            np.all(palette == expected_color, axis=1)
        )
        remap[matching_indices] = class_id
    return remap


def decode_label_array(
    raw_labels: np.ndarray,
    palette_remap: np.ndarray | None,
    label_path: Path,
) -> np.ndarray:
    labels = (
        palette_remap[raw_labels]
        if palette_remap is not None else raw_labels
    )
    if labels.size and (labels.min() < 0 or labels.max() >= len(CLASS_NAMES)):
        if palette_remap is not None:
            reason = "an indexed color outside the documented class palette"
        else:
            reason = f"a class ID outside 0-{len(CLASS_NAMES) - 1}"
        raise ValueError(f"{label_path} contains {reason}")
    return np.asarray(labels, dtype=np.uint8)


def load_label_data(
    label_path: Path,
    downsample: int,
    chunk_rows: int,
) -> tuple[np.ndarray, np.ndarray, int, int]:
    started = time.perf_counter()
    log(
        f"Loading indexed segmentation | path={label_path} | "
        f"bytes={file_size_or_unavailable(label_path)} | downsample={downsample} | "
        f"chunk_rows={chunk_rows}",
        logging.DEBUG,
    )
    try:
        with Image.open(label_path) as image:
            if image.mode not in {"P", "L", "I", "I;16"}:
                raise ValueError(
                    f"{label_path} must preserve indexed class IDs; mode={image.mode}"
                )
            palette_remap = palette_class_remap(image, label_path)
            width, height = image.size
            hard_counts = np.zeros(len(CLASS_NAMES), dtype=np.int64)
            chunk_count = math.ceil(height / chunk_rows)
            for top in range(0, height, chunk_rows):
                bottom = min(top + chunk_rows, height)
                chunk_index = top // chunk_rows + 1
                raw_chunk = np.asarray(
                    image.crop((0, top, width, bottom)), dtype=np.int64
                )
                chunk = decode_label_array(
                    raw_chunk, palette_remap, label_path
                )
                hard_counts += np.bincount(
                    chunk.reshape(-1), minlength=len(CLASS_NAMES)
                )[: len(CLASS_NAMES)]
                if (
                    chunk_index == 1
                    or chunk_index == chunk_count
                    or chunk_index % max(1, chunk_count // 10) == 0
                ):
                    update_status(
                        "Reading segmentation | "
                        f"file={label_path.name} | chunk={chunk_index}/{chunk_count} | "
                        f"rows={top}:{bottom}"
                    )
            analysis_width = math.ceil(width / downsample)
            analysis_height = math.ceil(height / downsample)
            resized = image.resize(
                (analysis_width, analysis_height), Image.Resampling.NEAREST
            )
            try:
                raw_labels = np.array(resized, dtype=np.int64, copy=True)
                labels = decode_label_array(
                    raw_labels, palette_remap, label_path
                )
            finally:
                resized.close()
    except UnidentifiedImageError as error:
        raise ValueError(f"Not a valid segmentation image: {label_path}") from error
    log(
        "Indexed segmentation loaded | "
        f"file={label_path.name} | native={width}x{height} | "
        f"analysis={analysis_width}x{analysis_height} | "
        f"hard_counts={dict(zip(CLASS_NAMES, hard_counts.tolist()))} | "
        f"elapsed_seconds={time.perf_counter() - started:.3f}",
        logging.DEBUG,
    )
    return labels, hard_counts, height, width


def hard_boundary_mask(labels: np.ndarray) -> np.ndarray:
    boundary = np.zeros(labels.shape, dtype=bool)
    horizontal = labels[:, 1:] != labels[:, :-1]
    vertical = labels[1:, :] != labels[:-1, :]
    boundary[:, 1:] |= horizontal
    boundary[:, :-1] |= horizontal
    boundary[1:, :] |= vertical
    boundary[:-1, :] |= vertical
    return boundary


def physical_skeleton_length(
    skeleton: np.ndarray,
    pixel_height_um: float,
    pixel_width_um: float,
) -> float:
    horizontal = np.count_nonzero(skeleton[:, :-1] & skeleton[:, 1:])
    vertical = np.count_nonzero(skeleton[:-1, :] & skeleton[1:, :])
    diagonal = (
        np.count_nonzero(skeleton[:-1, :-1] & skeleton[1:, 1:])
        + np.count_nonzero(skeleton[:-1, 1:] & skeleton[1:, :-1])
    )
    return float(
        horizontal * pixel_width_um
        + vertical * pixel_height_um
        + diagonal * math.hypot(pixel_width_um, pixel_height_um)
    )


def axial_orientation_statistics(
    angles_radians: np.ndarray,
    weights: np.ndarray | None = None,
) -> tuple[float | None, float | None]:
    angles = np.asarray(angles_radians, dtype=np.float64)
    valid = np.isfinite(angles)
    if weights is not None:
        weights = np.asarray(weights, dtype=np.float64)
        valid &= np.isfinite(weights) & (weights > 0)
    if not np.any(valid):
        return None, None
    doubled = 2.0 * angles[valid]
    if weights is None:
        vector = np.mean(np.exp(1j * doubled))
    else:
        valid_weights = weights[valid]
        vector = np.sum(valid_weights * np.exp(1j * doubled)) / np.sum(
            valid_weights
        )
    mean_angle = 0.5 * math.atan2(vector.imag, vector.real)
    return math.degrees(mean_angle), float(abs(vector))


def nearest_neighbor_statistics(
    points_um: np.ndarray,
) -> tuple[float | None, float | None, float | None]:
    if len(points_um) < 2:
        return None, None, None
    distances, _ = cKDTree(points_um).query(points_um, k=2)
    nearest = distances[:, 1]
    return (
        float(nearest.mean()),
        float(np.median(nearest)),
        float(nearest.max()),
    )


def curvature_from_signed_distance(
    mask: np.ndarray,
    pixel_height_um: float,
    pixel_width_um: float,
) -> tuple[np.ndarray, np.ndarray]:
    inside = accelerated_distance_transform_edt(
        mask, sampling=(pixel_height_um, pixel_width_um)
    )
    outside = accelerated_distance_transform_edt(
        ~mask, sampling=(pixel_height_um, pixel_width_um)
    )
    signed = outside - inside
    grad_y, grad_x = np.gradient(
        signed, pixel_height_um, pixel_width_um
    )
    norm = np.hypot(grad_y, grad_x)
    norm = np.maximum(norm, EPSILON)
    normal_y = grad_y / norm
    normal_x = grad_x / norm
    curvature = (
        np.gradient(normal_y, pixel_height_um, axis=0)
        + np.gradient(normal_x, pixel_width_um, axis=1)
    )
    return curvature.astype(np.float32), inside


def contour_length_um(
    mask: np.ndarray,
    pixel_height_um: float,
    pixel_width_um: float,
) -> float:
    total = 0.0
    padded_mask = np.pad(mask.astype(np.uint8), 1, mode="constant")
    for contour in measure.find_contours(padded_mask, 0.5):
        if len(contour) < 2:
            continue
        differences = np.diff(contour, axis=0)
        total += np.hypot(
            differences[:, 0] * pixel_height_um,
            differences[:, 1] * pixel_width_um,
        ).sum()
    return float(total)


def cpu_native_probability_chunk_stats(
    probabilities: np.ndarray,
    labels: np.ndarray,
    args: argparse.Namespace,
) -> NativeProbabilityChunkStats:
    class_count = len(CLASS_NAMES)
    expected_pixels = probabilities.sum(axis=(1, 2), dtype=np.float64)
    hard_confidence_sum = np.zeros(class_count, dtype=np.float64)
    hard_confidence_min = np.full(class_count, np.inf, dtype=np.float64)
    hard_confidence_count = np.zeros(class_count, dtype=np.int64)
    low_confidence_count = np.zeros(class_count, dtype=np.int64)
    for class_id in range(class_count):
        selected = probabilities[class_id][labels == class_id]
        if selected.size:
            hard_confidence_sum[class_id] = selected.sum(dtype=np.float64)
            hard_confidence_count[class_id] = selected.size
            hard_confidence_min[class_id] = float(selected.min())
            low_confidence_count[class_id] = np.count_nonzero(
                selected < args.low_confidence_threshold
            )

    sums = probabilities.sum(axis=0, keepdims=True)
    normalized = probabilities / np.maximum(sums, EPSILON)
    entropy = -np.sum(
        normalized * np.log(np.maximum(normalized, EPSILON)), axis=0
    ) / math.log(class_count)
    two_largest = np.partition(normalized, -2, axis=0)[-2:]
    margin = two_largest[1] - two_largest[0]
    # A strict comparison against every other class gives deterministic CPU/GPU
    # behavior when probability values tie. It also avoids a large integer
    # argpartition array for whole-slide chunks.
    bone_fibro_pair = np.minimum(
        normalized[0], normalized[1]
    ) > np.max(normalized[2:], axis=0)
    return NativeProbabilityChunkStats(
        expected_pixels=expected_pixels,
        hard_confidence_sum=hard_confidence_sum,
        hard_confidence_min=hard_confidence_min,
        hard_confidence_count=hard_confidence_count,
        low_confidence_count=low_confidence_count,
        entropy_sum=float(entropy.sum(dtype=np.float64)),
        high_entropy_count=int(
            np.count_nonzero(entropy >= args.high_entropy_threshold)
        ),
        low_margin_count=int(
            np.count_nonzero(margin <= args.uncertainty_margin)
        ),
        bone_fibro_ambiguity_count=int(
            np.count_nonzero(
                bone_fibro_pair & (margin <= args.uncertainty_margin)
            )
        ),
    )


def torch_native_probability_chunk_stats(
    probabilities: np.ndarray,
    labels: np.ndarray,
    args: argparse.Namespace,
) -> NativeProbabilityChunkStats | None:
    stage = "torch_probability_statistics"
    pixel_count = labels.size
    if not ACCELERATOR.should_use_torch(pixel_count, stage):
        return None
    torch = ACCELERATOR.torch
    try:
        with torch.inference_mode():
            probability_tensor = torch.as_tensor(
                np.ascontiguousarray(probabilities),
                device=ACCELERATOR.torch_device,
                dtype=torch.float32,
            )
            label_tensor = torch.as_tensor(
                np.ascontiguousarray(labels),
                device=ACCELERATOR.torch_device,
                dtype=torch.long,
            )
            class_count = len(CLASS_NAMES)
            expected_pixels = probability_tensor.sum(
                dim=(1, 2), dtype=torch.float64
            )
            hard_values = probability_tensor.gather(
                0, label_tensor.unsqueeze(0)
            ).squeeze(0)
            flat_labels = label_tensor.reshape(-1)
            flat_values = hard_values.reshape(-1)
            hard_confidence_count = torch.bincount(
                flat_labels, minlength=class_count
            )
            hard_confidence_sum = torch.bincount(
                flat_labels,
                weights=flat_values.to(torch.float64),
                minlength=class_count,
            )
            low_confidence_count = torch.bincount(
                flat_labels,
                weights=(
                    flat_values < args.low_confidence_threshold
                ).to(torch.float64),
                minlength=class_count,
            )
            hard_confidence_min = torch.full(
                (class_count,),
                float("inf"),
                dtype=torch.float64,
                device=ACCELERATOR.torch_device,
            )
            for class_id in range(class_count):
                selected = flat_values[flat_labels == class_id]
                if selected.numel():
                    hard_confidence_min[class_id] = selected.min().to(
                        torch.float64
                    )

            normalized = probability_tensor / probability_tensor.sum(
                dim=0, keepdim=True
            ).clamp_min(EPSILON)
            entropy = -(
                normalized
                * normalized.clamp_min(EPSILON).log()
            ).sum(dim=0) / math.log(class_count)
            top_values = torch.topk(
                normalized, k=2, dim=0, largest=True, sorted=True
            ).values
            margin = top_values[0] - top_values[1]
            low_margin = margin <= args.uncertainty_margin
            bone_fibro_pair = torch.minimum(
                normalized[0], normalized[1]
            ) > normalized[2:].amax(dim=0)
            result = NativeProbabilityChunkStats(
                expected_pixels=expected_pixels.cpu().numpy(),
                hard_confidence_sum=hard_confidence_sum.cpu().numpy(),
                hard_confidence_min=hard_confidence_min.cpu().numpy(),
                hard_confidence_count=hard_confidence_count.cpu().numpy(),
                low_confidence_count=np.rint(
                    low_confidence_count.cpu().numpy()
                ).astype(np.int64),
                entropy_sum=float(
                    entropy.sum(dtype=torch.float64).item()
                ),
                high_entropy_count=int(
                    (entropy >= args.high_entropy_threshold).sum().item()
                ),
                low_margin_count=int(low_margin.sum().item()),
                bone_fibro_ambiguity_count=int(
                    (bone_fibro_pair & low_margin).sum().item()
                ),
            )
            del (
                probability_tensor,
                label_tensor,
                hard_values,
                normalized,
                entropy,
                top_values,
            )
            if stage not in ACCELERATOR.reported_stages:
                ACCELERATOR.reported_stages.add(stage)
                log(
                    f"GPU stage active | stage={stage} | "
                    f"chunk_shape={probabilities.shape}",
                    logging.DEBUG,
                )
            return result
    except Exception as error:
        ACCELERATOR.disable_stage(stage, error)
        return None


def native_probability_chunk_stats(
    probabilities: np.ndarray,
    labels: np.ndarray,
    args: argparse.Namespace,
) -> NativeProbabilityChunkStats:
    gpu_result = torch_native_probability_chunk_stats(
        probabilities, labels, args
    )
    if gpu_result is not None:
        return gpu_result
    return cpu_native_probability_chunk_stats(probabilities, labels, args)


def cpu_uncertainty_chunk(
    probabilities: np.ndarray,
    uncertainty_margin: float,
) -> tuple[np.ndarray, np.ndarray]:
    normalized = probabilities / np.maximum(
        probabilities.sum(axis=0, keepdims=True), EPSILON
    )
    two_largest = np.partition(normalized, -2, axis=0)[-2:]
    margin = two_largest[1] - two_largest[0]
    uncertain = margin <= uncertainty_margin
    pair = np.minimum(normalized[0], normalized[1]) > np.max(
        normalized[2:], axis=0
    )
    return uncertain, pair & uncertain


def uncertainty_chunk(
    probabilities: np.ndarray,
    uncertainty_margin: float,
) -> tuple[np.ndarray, np.ndarray]:
    stage = "torch_analysis_uncertainty"
    pixel_count = probabilities.shape[1] * probabilities.shape[2]
    if ACCELERATOR.should_use_torch(pixel_count, stage):
        torch = ACCELERATOR.torch
        try:
            with torch.inference_mode():
                tensor = torch.as_tensor(
                    np.ascontiguousarray(probabilities),
                    device=ACCELERATOR.torch_device,
                    dtype=torch.float32,
                )
                normalized = tensor / tensor.sum(
                    dim=0, keepdim=True
                ).clamp_min(EPSILON)
                values = torch.topk(
                    normalized, k=2, dim=0, largest=True, sorted=True
                ).values
                uncertain = values[0] - values[1] <= uncertainty_margin
                pair = torch.minimum(
                    normalized[0], normalized[1]
                ) > normalized[2:].amax(dim=0)
                uncertain_cpu = uncertain.cpu().numpy()
                pair_cpu = (pair & uncertain).cpu().numpy()
                if stage not in ACCELERATOR.reported_stages:
                    ACCELERATOR.reported_stages.add(stage)
                    log(
                        f"GPU stage active | stage={stage} | "
                        f"chunk_shape={probabilities.shape}",
                        logging.DEBUG,
                    )
                return uncertain_cpu, pair_cpu
        except Exception as error:
            ACCELERATOR.disable_stage(stage, error)
    return cpu_uncertainty_chunk(probabilities, uncertainty_margin)


def probability_statistics(
    source: ProbabilitySource,
    label_path: Path,
    labels_analysis: np.ndarray,
    height: int,
    width: int,
    args: argparse.Namespace,
) -> tuple[
    dict[str, object],
    list[dict[str, object]],
    np.ndarray | None,
    np.ndarray | None,
]:
    if not source.available:
        log(
            f"{source.scan.name}: probability statistics skipped because no "
            "probability source is available",
            logging.WARNING,
        )
        return (
            {
                "Probability source": "Unavailable",
                "Mean normalized entropy": None,
                "High-entropy pixel fraction": None,
                "Low top-two margin fraction": None,
                "Bone-Fibrocartilage ambiguity fraction": None,
            },
            [
                {
                    "Expected pixels": None,
                    "Mean hard-region confidence": None,
                    "Minimum hard-region confidence": None,
                    "Low-confidence hard-pixel fraction": None,
                }
                for _ in CLASS_NAMES
            ],
            None,
            None,
        )

    started = time.perf_counter()
    log(
        f"{source.scan.name}: calculating native-resolution probability "
        f"statistics from {source.kind}"
    )
    expected_pixels = np.zeros(len(CLASS_NAMES), dtype=np.float64)
    hard_confidence_sum = np.zeros(len(CLASS_NAMES), dtype=np.float64)
    hard_confidence_min = np.full(len(CLASS_NAMES), np.inf, dtype=np.float64)
    hard_confidence_count = np.zeros(len(CLASS_NAMES), dtype=np.int64)
    low_confidence_count = np.zeros(len(CLASS_NAMES), dtype=np.int64)
    entropy_sum = 0.0
    high_entropy_count = 0
    low_margin_count = 0
    bone_fibro_ambiguity_count = 0
    total_pixels = height * width
    native_chunk_rows = args.chunk_rows
    if ACCELERATOR.torch_cuda_available:
        native_chunk_rows = ACCELERATOR.choose_chunk_rows(
            width, height, bytes_per_pixel=96
        )
        log(
            f"{source.scan.name}: CUDA probability chunk rows="
            f"{native_chunk_rows:,}",
            logging.DEBUG,
        )
    native_chunk_count = math.ceil(height / native_chunk_rows)

    with Image.open(label_path) as label_image:
        label_remap = palette_class_remap(label_image, label_path)
        for top in range(0, height, native_chunk_rows):
            bottom = min(top + native_chunk_rows, height)
            chunk_index = top // native_chunk_rows + 1
            if (
                chunk_index == 1
                or chunk_index == native_chunk_count
                or chunk_index % max(1, native_chunk_count // 10) == 0
            ):
                update_status(
                    f"{source.scan.name} | native probabilities | "
                    f"chunk {chunk_index}/{native_chunk_count}"
                )
            probabilities = np.clip(source.native_chunk(top, bottom), 0.0, 1.0)
            raw_labels = np.asarray(
                label_image.crop((0, top, width, bottom)), dtype=np.int64
            )
            labels = decode_label_array(
                raw_labels, label_remap, label_path
            )
            chunk_stats = native_probability_chunk_stats(
                probabilities, labels, args
            )
            expected_pixels += chunk_stats.expected_pixels
            hard_confidence_sum += chunk_stats.hard_confidence_sum
            hard_confidence_min = np.minimum(
                hard_confidence_min, chunk_stats.hard_confidence_min
            )
            hard_confidence_count += chunk_stats.hard_confidence_count
            low_confidence_count += chunk_stats.low_confidence_count
            entropy_sum += chunk_stats.entropy_sum
            high_entropy_count += chunk_stats.high_entropy_count
            low_margin_count += chunk_stats.low_margin_count
            bone_fibro_ambiguity_count += (
                chunk_stats.bone_fibro_ambiguity_count
            )
            del probabilities, labels, chunk_stats

    analysis_height, analysis_width = labels_analysis.shape
    uncertain_mask = np.zeros(labels_analysis.shape, dtype=bool)
    bone_fibro_mask = np.zeros(labels_analysis.shape, dtype=bool)
    analysis_chunk_rows = args.chunk_rows
    if ACCELERATOR.torch_cuda_available:
        analysis_chunk_rows = ACCELERATOR.choose_chunk_rows(
            analysis_width, analysis_height, bytes_per_pixel=72
        )
    analysis_chunk_count = math.ceil(analysis_height / analysis_chunk_rows)
    for top in range(0, analysis_height, analysis_chunk_rows):
        bottom = min(top + analysis_chunk_rows, analysis_height)
        chunk_index = top // analysis_chunk_rows + 1
        if (
            chunk_index == 1
            or chunk_index == analysis_chunk_count
            or chunk_index % max(1, analysis_chunk_count // 10) == 0
        ):
            update_status(
                f"{source.scan.name} | analysis-scale uncertainty | "
                f"chunk {chunk_index}/{analysis_chunk_count}"
            )
        probabilities = np.clip(
            source.analysis_chunk(
                top, bottom, analysis_height, analysis_width
            ),
            0.0,
            1.0,
        )
        uncertain_chunk, bone_fibro_chunk = uncertainty_chunk(
            probabilities, args.uncertainty_margin
        )
        uncertain_mask[top:bottom] = uncertain_chunk
        bone_fibro_mask[top:bottom] = bone_fibro_chunk
        del probabilities, uncertain_chunk, bone_fibro_chunk

    class_rows = []
    for class_id in range(len(CLASS_NAMES)):
        count = int(hard_confidence_count[class_id])
        class_rows.append(
            {
                "Expected pixels": float(expected_pixels[class_id]),
                "Mean hard-region confidence": (
                    float(hard_confidence_sum[class_id] / count)
                    if count else None
                ),
                "Minimum hard-region confidence": (
                    float(hard_confidence_min[class_id]) if count else None
                ),
                "Low-confidence hard-pixel fraction": (
                    float(low_confidence_count[class_id] / count)
                    if count else None
                ),
            }
        )
    scan_metrics = {
        "Probability source": source.kind,
        "Mean normalized entropy": float(entropy_sum / total_pixels),
        "High-entropy pixel fraction": float(
            high_entropy_count / total_pixels
        ),
        "Low top-two margin fraction": float(
            low_margin_count / total_pixels
        ),
        "Bone-Fibrocartilage ambiguity fraction": float(
            bone_fibro_ambiguity_count / total_pixels
        ),
    }
    log(
        "Probability statistics complete | "
        f"scan={source.scan.name} | source={source.kind} | "
        f"mean_entropy={scan_metrics['Mean normalized entropy']:.6f} | "
        f"high_entropy_fraction={scan_metrics['High-entropy pixel fraction']:.6f} | "
        f"low_margin_fraction={scan_metrics['Low top-two margin fraction']:.6f} | "
        f"bone_fibro_ambiguity_fraction="
        f"{scan_metrics['Bone-Fibrocartilage ambiguity fraction']:.6f} | "
        f"elapsed_seconds={time.perf_counter() - started:.3f}"
    )
    # Probability chunks use the PyTorch caching allocator. Return those
    # cached blocks before CuPy begins morphology and distance-transform work.
    ACCELERATOR.release_memory()
    return (
        scan_metrics,
        class_rows,
        uncertain_mask,
        bone_fibro_mask,
    )


def component_measurements(
    scan_name: str,
    class_name: str,
    mask: np.ndarray,
    class_probability: np.ndarray | None,
    curvature: np.ndarray | None,
    pixel_height_um: float,
    pixel_width_um: float,
    low_confidence_threshold: float,
    detail_limit: int,
) -> tuple[
    dict[str, object],
    list[dict[str, object]],
    np.ndarray,
    list[measure._regionprops.RegionProperties],
    int,
]:
    component_labels, component_count = accelerated_label(
        mask, structure=np.ones((3, 3), dtype=np.uint8)
    )
    properties = measure.regionprops(component_labels)
    sorted_properties = sorted(properties, key=lambda item: item.area, reverse=True)
    detail_ids = {prop.label for prop in sorted_properties[:detail_limit]}
    detail_rows: list[dict[str, object]] = []
    areas_um2: list[float] = []
    perimeters_um: list[float] = []
    circularities: list[float] = []
    solidities: list[float] = []
    convexities: list[float] = []
    roughnesses: list[float] = []
    eccentricities: list[float] = []
    aspect_ratios: list[float] = []
    orientations: list[float] = []
    curvatures: list[float] = []
    boundary_touching_count = 0
    boundary_touching_area = 0.0
    image_height, image_width = mask.shape

    for prop in properties:
        local_mask = prop.image
        area_um2 = float(prop.area * pixel_height_um * pixel_width_um)
        perimeter_um = contour_length_um(
            local_mask, pixel_height_um, pixel_width_um
        )
        convex_mask = prop.image_convex
        convex_perimeter_um = contour_length_um(
            convex_mask, pixel_height_um, pixel_width_um
        )
        circularity = (
            float(4.0 * math.pi * area_um2 / perimeter_um ** 2)
            if perimeter_um > 0 else None
        )
        convexity = (
            float(convex_perimeter_um / perimeter_um)
            if perimeter_um > 0 else None
        )
        roughness = (
            float(perimeter_um / convex_perimeter_um)
            if convex_perimeter_um > 0 else None
        )
        physical_coordinates = prop.coords.astype(np.float64)
        physical_coordinates[:, 0] *= pixel_height_um
        physical_coordinates[:, 1] *= pixel_width_um
        centered_coordinates = physical_coordinates - physical_coordinates.mean(
            axis=0, keepdims=True
        )
        covariance = (
            centered_coordinates.T @ centered_coordinates
        ) / max(len(centered_coordinates), 1)
        eigenvalues, eigenvectors = np.linalg.eigh(covariance)
        eigenvalues = np.maximum(eigenvalues, 0.0)
        major_um = float(4.0 * math.sqrt(eigenvalues[-1]))
        minor_um = float(4.0 * math.sqrt(eigenvalues[0]))
        aspect_ratio = major_um / minor_um if minor_um > 0 else None
        major_vector = eigenvectors[:, -1]
        orientation_radians = math.atan2(major_vector[0], major_vector[1])
        orientation_deg = math.degrees(orientation_radians)
        physical_eccentricity = (
            math.sqrt(max(0.0, 1.0 - eigenvalues[0] / eigenvalues[-1]))
            if eigenvalues[-1] > 0 else 0.0
        )
        min_row, min_col, max_row, max_col = prop.bbox
        touches_border = (
            min_row == 0 or min_col == 0
            or max_row == image_height or max_col == image_width
        )
        if touches_border:
            boundary_touching_count += 1
            boundary_touching_area += area_um2

        region_curvature = None
        if curvature is not None:
            local_curvature = curvature[prop.slice]
            region_boundary = local_mask & ~accelerated_binary_erosion(local_mask)
            values = np.abs(local_curvature[region_boundary])
            if values.size:
                region_curvature = float(np.mean(values))

        mean_confidence = None
        min_confidence = None
        low_confidence_fraction = None
        if class_probability is not None:
            values = class_probability[prop.slice][local_mask]
            if values.size:
                mean_confidence = float(values.mean())
                min_confidence = float(values.min())
                low_confidence_fraction = float(
                    np.count_nonzero(values < low_confidence_threshold)
                    / values.size
                )

        areas_um2.append(area_um2)
        perimeters_um.append(perimeter_um)
        if circularity is not None:
            circularities.append(circularity)
        solidities.append(float(prop.solidity))
        if convexity is not None:
            convexities.append(convexity)
        if roughness is not None:
            roughnesses.append(roughness)
        eccentricities.append(physical_eccentricity)
        if aspect_ratio is not None:
            aspect_ratios.append(aspect_ratio)
        orientations.append(orientation_radians)
        if region_curvature is not None:
            curvatures.append(region_curvature)

        if prop.label in detail_ids:
            detail_rows.append(
                {
                    "Scan": scan_name,
                    "Class": class_name,
                    "Region ID": int(prop.label),
                    "Area (µm²)": area_um2,
                    "Perimeter (µm)": perimeter_um,
                    "Perimeter / area (1/µm)": (
                        perimeter_um / area_um2 if area_um2 > 0 else None
                    ),
                    "Equivalent diameter (µm)": float(
                        prop.equivalent_diameter_area
                        * math.sqrt(pixel_height_um * pixel_width_um)
                    ),
                    "Major axis (µm)": major_um,
                    "Minor axis (µm)": minor_um,
                    "Aspect ratio": aspect_ratio,
                    "Elongation": (
                        1.0 - minor_um / major_um if major_um > 0 else None
                    ),
                    "Eccentricity": physical_eccentricity,
                    "Orientation (degrees)": orientation_deg,
                    "Circularity": circularity,
                    "Solidity": float(prop.solidity),
                    "Convexity": convexity,
                    "Boundary roughness": roughness,
                    "Extent": float(prop.extent),
                    "Centroid X (µm)": float(prop.centroid[1] * pixel_width_um),
                    "Centroid Y (µm)": float(prop.centroid[0] * pixel_height_um),
                    "Touches image boundary": touches_border,
                    "Mean absolute curvature (1/µm)": region_curvature,
                    "Mean confidence": mean_confidence,
                    "Minimum confidence": min_confidence,
                    "Low-confidence pixel fraction": low_confidence_fraction,
                    "RGB variation within region": None,
                    "Intensity entropy within region (bits)": None,
                    "Gradient mean within region": None,
                    "Local entropy mean within region (bits)": None,
                    "Structure coherence within region": None,
                    "Texture orientation within region (degrees)": None,
                    "Texture-orientation regularity within region": None,
                    "Hessian ridge mean within region": None,
                }
            )

    total_component_area = float(np.sum(areas_um2))
    orientation_mean, orientation_regularity = axial_orientation_statistics(
        np.asarray(orientations), np.asarray(areas_um2)
    )
    centroids_um = np.asarray(
        [
            [prop.centroid[0] * pixel_height_um, prop.centroid[1] * pixel_width_um]
            for prop in properties
        ],
        dtype=np.float64,
    ).reshape((-1, 2))
    nearest_mean, nearest_median, nearest_max = nearest_neighbor_statistics(
        centroids_um
    )
    image_area_um2 = (
        image_height * pixel_height_um * image_width * pixel_width_um
    )
    density_per_um2 = component_count / image_area_um2 if image_area_um2 else 0
    expected_random_nn = (
        0.5 / math.sqrt(density_per_um2)
        if component_count >= 2 and density_per_um2 > 0 else None
    )
    clark_evans = (
        nearest_mean / expected_random_nn
        if nearest_mean is not None and expected_random_nn else None
    )
    summary = {
        "Region count": int(component_count),
        "Region density (per mm² image)": (
            component_count / (image_area_um2 / 1_000_000.0)
            if image_area_um2 else None
        ),
        "Largest region area (µm²)": max(areas_um2, default=0.0),
        "Largest-region area fraction": (
            max(areas_um2, default=0.0) / total_component_area
            if total_component_area else None
        ),
        "Fragmentation fraction": (
            1.0 - max(areas_um2, default=0.0) / total_component_area
            if total_component_area else None
        ),
        "Region area mean (µm²)": safe_stat(areas_um2, "mean"),
        "Region area median (µm²)": safe_stat(areas_um2, "median"),
        "Region area SD (µm²)": safe_stat(areas_um2, "std"),
        "Region area P10 (µm²)": safe_stat(areas_um2, "p10"),
        "Region area P90 (µm²)": safe_stat(areas_um2, "p90"),
        "Total perimeter (µm)": float(np.sum(perimeters_um)),
        "Mean circularity": safe_stat(circularities, "mean"),
        "Mean solidity": safe_stat(solidities, "mean"),
        "Mean convexity": safe_stat(convexities, "mean"),
        "Mean boundary roughness": safe_stat(roughnesses, "mean"),
        "Mean eccentricity": safe_stat(eccentricities, "mean"),
        "Mean aspect ratio": safe_stat(aspect_ratios, "mean"),
        "Mean absolute curvature (1/µm)": safe_stat(curvatures, "mean"),
        "Mean orientation (degrees)": orientation_mean,
        "Region-orientation regularity": orientation_regularity,
        "Boundary-touching region fraction": (
            boundary_touching_count / component_count if component_count else None
        ),
        "Boundary-touching area fraction": (
            boundary_touching_area / total_component_area
            if total_component_area else None
        ),
        "Same-class nearest-neighbor mean (µm)": nearest_mean,
        "Same-class nearest-neighbor median (µm)": nearest_median,
        "Same-class nearest-neighbor max (µm)": nearest_max,
        "Clark-Evans clustering index": clark_evans,
    }
    omitted = max(0, component_count - len(detail_rows))
    return summary, detail_rows, component_labels, properties, omitted


def pore_measurements(
    scan_name: str,
    class_name: str,
    mask: np.ndarray,
    pixel_height_um: float,
    pixel_width_um: float,
    detail_limit: int,
) -> tuple[dict[str, object], list[dict[str, object]], np.ndarray, int]:
    filled = accelerated_binary_fill_holes(mask)
    pores = filled & ~mask
    pore_labels, pore_count = accelerated_label(
        pores, structure=np.ones((3, 3), dtype=np.uint8)
    )
    properties = measure.regionprops(pore_labels)
    sorted_properties = sorted(properties, key=lambda item: item.area, reverse=True)
    detail_ids = {prop.label for prop in sorted_properties[:detail_limit]}
    pore_areas = np.asarray(
        [prop.area * pixel_height_um * pixel_width_um for prop in properties],
        dtype=np.float64,
    )
    pore_diameters = np.asarray(
        [
            prop.equivalent_diameter_area
            * math.sqrt(pixel_height_um * pixel_width_um)
            for prop in properties
        ],
        dtype=np.float64,
    )
    centroids_um = np.asarray(
        [
            [prop.centroid[0] * pixel_height_um, prop.centroid[1] * pixel_width_um]
            for prop in properties
        ],
        dtype=np.float64,
    ).reshape((-1, 2))
    nearest_mean, nearest_median, nearest_max = nearest_neighbor_statistics(
        centroids_um
    )
    pore_area = float(pore_areas.sum())
    tissue_and_pore_area = float(
        np.count_nonzero(filled) * pixel_height_um * pixel_width_um
    )
    tissue_area = float(
        np.count_nonzero(mask) * pixel_height_um * pixel_width_um
    )
    detail_rows = []
    for prop in properties:
        if prop.label not in detail_ids:
            continue
        detail_rows.append(
            {
                "Scan": scan_name,
                "Class": class_name,
                "Pore ID": int(prop.label),
                "Area (µm²)": float(
                    prop.area * pixel_height_um * pixel_width_um
                ),
                "Equivalent diameter (µm)": float(
                    prop.equivalent_diameter_area
                    * math.sqrt(pixel_height_um * pixel_width_um)
                ),
                "Centroid X (µm)": float(prop.centroid[1] * pixel_width_um),
                "Centroid Y (µm)": float(prop.centroid[0] * pixel_height_um),
                "Eccentricity": float(prop.eccentricity),
                "Solidity": float(prop.solidity),
            }
        )
    summary = {
        "Closed-hole count": int(pore_count),
        "Euler number": int(measure.euler_number(mask, connectivity=2)),
        "Enclosed pore area (µm²)": pore_area,
        "Enclosed porosity fraction": (
            pore_area / tissue_and_pore_area if tissue_and_pore_area else None
        ),
        "Pore density (per mm² tissue)": (
            pore_count / (tissue_area / 1_000_000.0) if tissue_area else None
        ),
        "Pore area mean (µm²)": safe_stat(pore_areas, "mean"),
        "Pore area median (µm²)": safe_stat(pore_areas, "median"),
        "Pore area P90 (µm²)": safe_stat(pore_areas, "p90"),
        "Pore diameter mean (µm)": safe_stat(pore_diameters, "mean"),
        "Pore diameter median (µm)": safe_stat(pore_diameters, "median"),
        "Pore diameter P90 (µm)": safe_stat(pore_diameters, "p90"),
        "Pore nearest-neighbor mean (µm)": nearest_mean,
        "Pore nearest-neighbor median (µm)": nearest_median,
        "Pore nearest-neighbor max (µm)": nearest_max,
    }
    omitted = max(0, pore_count - len(detail_rows))
    return summary, detail_rows, pores, omitted


def skeleton_measurements(
    mask: np.ndarray,
    pores: np.ndarray,
    inside_distance: np.ndarray | None,
    pixel_height_um: float,
    pixel_width_um: float,
    skip_skeleton: bool,
) -> dict[str, object]:
    if skip_skeleton or not np.any(mask):
        return {
            "Skeleton length (µm)": None,
            "Skeleton length density (µm/mm² tissue)": None,
            "Skeleton endpoint count": None,
            "Skeleton junction count": None,
            "Local thickness mean (µm)": None,
            "Local thickness median (µm)": None,
            "Local thickness P90 (µm)": None,
            "Local thickness maximum (µm)": None,
            "Internal-void skeleton length (µm)": None,
            "Internal-void skeleton density (µm/mm² tissue)": None,
        }
    skeleton = morphology.skeletonize(mask)
    length_um = physical_skeleton_length(
        skeleton, pixel_height_um, pixel_width_um
    )
    neighbor_count = accelerated_convolve(
        skeleton.astype(np.uint8),
        np.ones((3, 3), dtype=np.uint8),
        mode="constant",
        cval=0,
    ) - skeleton.astype(np.uint8)
    endpoint_count = int(np.count_nonzero(skeleton & (neighbor_count == 1)))
    junction_pixels = skeleton & (neighbor_count >= 3)
    _, junction_count = accelerated_label(
        junction_pixels, structure=np.ones((3, 3), dtype=np.uint8)
    )
    if inside_distance is None:
        inside_distance = accelerated_distance_transform_edt(
            mask, sampling=(pixel_height_um, pixel_width_um)
        )
    thickness = 2.0 * inside_distance[skeleton]
    void_skeleton = morphology.skeletonize(pores) if np.any(pores) else pores
    void_length_um = physical_skeleton_length(
        void_skeleton, pixel_height_um, pixel_width_um
    )
    tissue_area_mm2 = (
        np.count_nonzero(mask)
        * pixel_height_um
        * pixel_width_um
        / 1_000_000.0
    )
    return {
        "Skeleton length (µm)": length_um,
        "Skeleton length density (µm/mm² tissue)": (
            length_um / tissue_area_mm2 if tissue_area_mm2 else None
        ),
        "Skeleton endpoint count": endpoint_count,
        "Skeleton junction count": int(junction_count),
        "Local thickness mean (µm)": safe_stat(thickness, "mean"),
        "Local thickness median (µm)": safe_stat(thickness, "median"),
        "Local thickness P90 (µm)": safe_stat(thickness, "p90"),
        "Local thickness maximum (µm)": safe_stat(thickness, "max"),
        "Internal-void skeleton length (µm)": void_length_um,
        "Internal-void skeleton density (µm/mm² tissue)": (
            void_length_um / tissue_area_mm2 if tissue_area_mm2 else None
        ),
    }


def analyze_classes(
    scan: ScanInputs,
    labels: np.ndarray,
    hard_counts: np.ndarray,
    probability_source: ProbabilitySource,
    probability_class_rows: list[dict[str, object]],
    args: argparse.Namespace,
) -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    dict[int, np.ndarray],
    dict[tuple[int, int], float | None],
    int,
    int,
]:
    downsample = args.analysis_downsample
    pixel_height_um = args.pixel_height_um * downsample
    pixel_width_um = args.pixel_width_um * downsample
    pixel_area_um2 = args.pixel_height_um * args.pixel_width_um
    total_pixels = int(hard_counts.sum())
    non_background_pixels = int(hard_counts[:-1].sum())
    expected_non_background = sum(
        float(row["Expected pixels"])
        for row in probability_class_rows[:-1]
        if row["Expected pixels"] is not None
    )
    class_rows: list[dict[str, object]] = []
    region_rows: list[dict[str, object]] = []
    pore_rows: list[dict[str, object]] = []
    centroids_by_class: dict[int, np.ndarray] = {}
    interface_curvature: dict[tuple[int, int], float | None] = {}
    region_rows_omitted = 0
    pore_rows_omitted = 0

    for class_id, class_name in enumerate(CLASS_NAMES):
        class_started = time.perf_counter()
        log(f"{scan.name}: analyzing class {class_id + 1}/{len(CLASS_NAMES)} ({class_name})")
        update_status(
            f"{scan.name} | {class_name} | preparing mask and probability map"
        )
        mask = labels == class_id
        class_probability = (
            probability_source.analysis_class(
                class_id, labels.shape[0], labels.shape[1], args.chunk_rows
            )
            if probability_source.available else None
        )
        curvature = None
        inside_distance = None
        if np.any(mask):
            update_status(f"{scan.name} | {class_name} | curvature and distances")
            if args.skip_curvature:
                if not args.skip_skeleton:
                    inside_distance = accelerated_distance_transform_edt(
                        mask, sampling=(pixel_height_um, pixel_width_um)
                    )
            else:
                curvature, inside_distance = curvature_from_signed_distance(
                    mask, pixel_height_um, pixel_width_um
                )
                class_boundary = mask & ~accelerated_binary_erosion(mask)
                for target_id in range(len(CLASS_NAMES)):
                    if target_id == class_id:
                        continue
                    target = labels == target_id
                    interface_side = class_boundary & accelerated_binary_dilation(target)
                    values = np.abs(curvature[interface_side])
                    interface_curvature[(class_id, target_id)] = (
                        float(values.mean()) if values.size else None
                    )
        remaining_regions = max(0, args.max_region_rows - len(region_rows))
        update_status(f"{scan.name} | {class_name} | connected-region morphology")
        morphology_summary, details, component_labels, properties, omitted = (
            component_measurements(
                scan.name,
                class_name,
                mask,
                class_probability,
                curvature,
                pixel_height_um,
                pixel_width_um,
                args.low_confidence_threshold,
                remaining_regions,
            )
        )
        region_rows.extend(details)
        region_rows_omitted += omitted

        remaining_pores = max(0, args.max_pore_rows - len(pore_rows))
        update_status(f"{scan.name} | {class_name} | closed holes and porosity")
        pore_summary, pore_details, pores, pore_omitted = pore_measurements(
            scan.name,
            class_name,
            mask,
            pixel_height_um,
            pixel_width_um,
            remaining_pores,
        )
        pore_rows.extend(pore_details)
        pore_rows_omitted += pore_omitted
        update_status(f"{scan.name} | {class_name} | skeleton and local thickness")
        skeleton_summary = skeleton_measurements(
            mask,
            pores,
            inside_distance,
            pixel_height_um,
            pixel_width_um,
            args.skip_skeleton,
        )

        centroids = np.asarray(
            [
                [
                    prop.centroid[0] * pixel_height_um,
                    prop.centroid[1] * pixel_width_um,
                ]
                for prop in properties
            ],
            dtype=np.float64,
        ).reshape((-1, 2))
        centroids_by_class[class_id] = centroids
        center = np.asarray(
            [
                labels.shape[0] * pixel_height_um / 2.0,
                labels.shape[1] * pixel_width_um / 2.0,
            ]
        )
        radial_distances = (
            np.linalg.norm(centroids - center, axis=1)
            if centroids.size else np.asarray([], dtype=np.float64)
        )
        if centroids.size:
            angles = np.arctan2(
                centroids[:, 0] - center[0], centroids[:, 1] - center[1]
            )
            directional_vector = np.mean(np.exp(1j * angles))
            mean_direction = math.degrees(
                math.atan2(directional_vector.imag, directional_vector.real)
            )
            directional_concentration = float(abs(directional_vector))
        else:
            mean_direction = None
            directional_concentration = None

        hard_pixels = int(hard_counts[class_id])
        hard_fraction_image = hard_pixels / total_pixels if total_pixels else None
        if class_name == "Background":
            hard_fraction_reference = hard_fraction_image
        else:
            hard_fraction_reference = (
                hard_pixels / non_background_pixels
                if non_background_pixels else None
            )
        probability_metrics = probability_class_rows[class_id]
        expected_pixels = probability_metrics["Expected pixels"]
        class_row: dict[str, object] = {
            "Scan": scan.name,
            "Class ID": class_id,
            "Class": class_name,
            "Hard pixels": hard_pixels,
            "Hard area (µm²)": hard_pixels * pixel_area_um2,
            "Hard area (mm²)": hard_pixels * pixel_area_um2 / 1_000_000.0,
            "Hard image-area fraction": hard_fraction_image,
            "Hard reference-area fraction": hard_fraction_reference,
            "Expected pixels": expected_pixels,
            "Expected area (µm²)": (
                float(expected_pixels) * pixel_area_um2
                if expected_pixels is not None else None
            ),
            "Expected area (mm²)": (
                float(expected_pixels) * pixel_area_um2 / 1_000_000.0
                if expected_pixels is not None else None
            ),
            "Confidence-weighted area (mm²)": (
                float(expected_pixels) * pixel_area_um2 / 1_000_000.0
                if expected_pixels is not None else None
            ),
            "Expected non-background fraction": (
                float(expected_pixels) / expected_non_background
                if class_name != "Background"
                and expected_pixels is not None
                and expected_non_background else None
            ),
            "Analysis-scale area (µm²)": float(
                np.count_nonzero(mask) * pixel_height_um * pixel_width_um
            ),
            "Radial centroid distance mean (µm)": safe_stat(
                radial_distances, "mean"
            ),
            "Radial centroid distance median (µm)": safe_stat(
                radial_distances, "median"
            ),
            "Mean centroid direction (degrees)": mean_direction,
            "Centroid directional concentration": directional_concentration,
            **probability_metrics,
            **morphology_summary,
            **pore_summary,
            **skeleton_summary,
        }
        class_rows.append(class_row)
        log(
            "Class analysis complete | "
            f"scan={scan.name} | class={class_name} | hard_pixels={hard_pixels:,} | "
            f"regions={morphology_summary['Region count']:,} | "
            f"closed_holes={pore_summary['Closed-hole count']:,} | "
            f"retained_region_rows={len(details):,} | "
            f"retained_pore_rows={len(pore_details):,} | "
            f"elapsed_seconds={time.perf_counter() - class_started:.3f}"
        )
        del mask, component_labels, pores, curvature, inside_distance
        if class_probability is not None:
            del class_probability
        gc.collect()

    return (
        class_rows,
        region_rows,
        pore_rows,
        centroids_by_class,
        interface_curvature,
        region_rows_omitted,
        pore_rows_omitted,
    )


def pair_contact_measurements(
    scan_name: str,
    labels: np.ndarray,
    interface_curvature: dict[tuple[int, int], float | None],
    uncertain_mask: np.ndarray | None,
    bone_fibro_mask: np.ndarray | None,
    pixel_height_um: float,
    pixel_width_um: float,
) -> tuple[list[dict[str, object]], dict[tuple[int, int], float]]:
    contact_lengths: dict[tuple[int, int], float] = defaultdict(float)
    class_contact_totals = defaultdict(float)
    horizontal_left = labels[:, :-1]
    horizontal_right = labels[:, 1:]
    vertical_top = labels[:-1, :]
    vertical_bottom = labels[1:, :]
    for first in range(len(CLASS_NAMES)):
        for second in range(first + 1, len(CLASS_NAMES)):
            horizontal = (
                ((horizontal_left == first) & (horizontal_right == second))
                | ((horizontal_left == second) & (horizontal_right == first))
            )
            vertical = (
                ((vertical_top == first) & (vertical_bottom == second))
                | ((vertical_top == second) & (vertical_bottom == first))
            )
            length_um = float(
                np.count_nonzero(horizontal) * pixel_height_um
                + np.count_nonzero(vertical) * pixel_width_um
            )
            contact_lengths[(first, second)] = length_um
            class_contact_totals[first] += length_um
            class_contact_totals[second] += length_um

    rows: list[dict[str, object]] = []
    pixel_area_um2 = pixel_height_um * pixel_width_um
    for first in range(len(CLASS_NAMES)):
        for second in range(first + 1, len(CLASS_NAMES)):
            pair = (first, second)
            length_um = contact_lengths[pair]
            pair_boundary = np.zeros(labels.shape, dtype=bool)
            horizontal = (
                ((horizontal_left == first) & (horizontal_right == second))
                | ((horizontal_left == second) & (horizontal_right == first))
            )
            vertical = (
                ((vertical_top == first) & (vertical_bottom == second))
                | ((vertical_top == second) & (vertical_bottom == first))
            )
            pair_boundary[:, :-1] |= horizontal
            pair_boundary[:, 1:] |= horizontal
            pair_boundary[:-1, :] |= vertical
            pair_boundary[1:, :] |= vertical
            _, segment_count = accelerated_label(
                pair_boundary, structure=np.ones((3, 3), dtype=np.uint8)
            )
            coordinates = np.argwhere(pair_boundary)
            if len(coordinates):
                row_span = (
                    coordinates[:, 0].max() - coordinates[:, 0].min() + 1
                ) * pixel_height_um
                col_span = (
                    coordinates[:, 1].max() - coordinates[:, 1].min() + 1
                ) * pixel_width_um
                bounding_diagonal = math.hypot(row_span, col_span)
                roughness_proxy = (
                    length_um / bounding_diagonal if bounding_diagonal else None
                )
            else:
                roughness_proxy = None
            curvature_values = [
                value
                for value in (
                    interface_curvature.get((first, second)),
                    interface_curvature.get((second, first)),
                )
                if value is not None
            ]
            uncertain_fraction = (
                float(np.count_nonzero(pair_boundary & uncertain_mask))
                / np.count_nonzero(pair_boundary)
                if uncertain_mask is not None and np.any(pair_boundary)
                else None
            )
            blend_area_um2 = None
            blend_width_um = None
            if pair == (0, 1) and bone_fibro_mask is not None:
                blend_area_um2 = float(
                    np.count_nonzero(bone_fibro_mask) * pixel_area_um2
                )
                blend_width_um = (
                    blend_area_um2 / length_um if length_um > 0 else None
                )
            rows.append(
                {
                    "Scan": scan_name,
                    "Class A": CLASS_NAMES[first],
                    "Class B": CLASS_NAMES[second],
                    "Adjacent": length_um > 0,
                    "Interface length (µm)": length_um,
                    "Interface segment count": int(segment_count),
                    "Class A contact fraction": (
                        length_um / class_contact_totals[first]
                        if class_contact_totals[first] else None
                    ),
                    "Class B contact fraction": (
                        length_um / class_contact_totals[second]
                        if class_contact_totals[second] else None
                    ),
                    "Interface roughness proxy": roughness_proxy,
                    "Mean absolute interface curvature (1/µm)": (
                        float(np.mean(curvature_values))
                        if curvature_values else None
                    ),
                    "Uncertain interface-pixel fraction": uncertain_fraction,
                    "Bone-Fibrocartilage ambiguous area (µm²)": blend_area_um2,
                    "Bone-Fibrocartilage transition-width proxy (µm)": blend_width_um,
                }
            )
            del pair_boundary, coordinates
    return rows, contact_lengths


def spatial_measurements(
    scan_name: str,
    labels: np.ndarray,
    centroids: dict[int, np.ndarray],
    contact_lengths: dict[tuple[int, int], float],
    pixel_height_um: float,
    pixel_width_um: float,
    proximity_um: float,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for target_id, target_name in enumerate(CLASS_NAMES):
        target_mask = labels == target_id
        if np.any(target_mask):
            distance = accelerated_distance_transform_edt(
                ~target_mask, sampling=(pixel_height_um, pixel_width_um)
            )
        else:
            distance = None
        target_centroids = centroids.get(
            target_id, np.empty((0, 2), dtype=np.float64)
        )
        target_tree = cKDTree(target_centroids) if len(target_centroids) else None
        for source_id, source_name in enumerate(CLASS_NAMES):
            if source_id == target_id:
                continue
            source_mask = labels == source_id
            values = distance[source_mask] if distance is not None else np.asarray([])
            source_centroids = centroids.get(
                source_id, np.empty((0, 2), dtype=np.float64)
            )
            if target_tree is not None and len(source_centroids):
                centroid_distances, _ = target_tree.query(source_centroids, k=1)
            else:
                centroid_distances = np.asarray([], dtype=np.float64)
            unordered = tuple(sorted((source_id, target_id)))
            rows.append(
                {
                    "Scan": scan_name,
                    "Source class": source_name,
                    "Target class": target_name,
                    "Classes adjacent": contact_lengths.get(unordered, 0.0) > 0,
                    "Source pixels": int(np.count_nonzero(source_mask)),
                    "Pixel-to-target distance minimum (µm)": safe_stat(
                        values, "min"
                    ),
                    "Pixel-to-target distance mean (µm)": safe_stat(
                        values, "mean"
                    ),
                    "Pixel-to-target distance median (µm)": safe_stat(
                        values, "median"
                    ),
                    "Pixel-to-target distance P90 (µm)": safe_stat(
                        values, "p90"
                    ),
                    f"Source area within {proximity_um:g} µm fraction": (
                        float(np.count_nonzero(values <= proximity_um) / values.size)
                        if values.size else None
                    ),
                    "Region-centroid nearest distance mean (µm)": safe_stat(
                        centroid_distances, "mean"
                    ),
                    "Region-centroid nearest distance median (µm)": safe_stat(
                        centroid_distances, "median"
                    ),
                    "Region-centroid nearest distance maximum (µm)": safe_stat(
                        centroid_distances, "max"
                    ),
                }
            )
        del distance
        gc.collect()
    return rows


def distance_bands_from_bone(
    scan_name: str,
    labels: np.ndarray,
    pixel_height_um: float,
    pixel_width_um: float,
    band_size_um: float,
) -> list[dict[str, object]]:
    bone_mask = labels == 0
    if not np.any(bone_mask):
        return []
    distance = accelerated_distance_transform_edt(
        ~bone_mask, sampling=(pixel_height_um, pixel_width_um)
    )
    band_ids = np.floor(distance / band_size_um).astype(np.int32)
    rows = []
    pixel_area_um2 = pixel_height_um * pixel_width_um
    for class_id, class_name in enumerate(CLASS_NAMES):
        selected_bands = band_ids[labels == class_id]
        if not selected_bands.size:
            continue
        counts = np.bincount(selected_bands)
        total = int(counts.sum())
        for band_id, count in enumerate(counts):
            if count == 0:
                continue
            rows.append(
                {
                    "Scan": scan_name,
                    "Class": class_name,
                    "Distance lower bound (µm)": band_id * band_size_um,
                    "Distance upper bound (µm)": (band_id + 1) * band_size_um,
                    "Pixels": int(count),
                    "Area (µm²)": float(count * pixel_area_um2),
                    "Class-area fraction in band": float(count / total),
                }
            )
    return rows


class PillowRegionSource:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.image: Image.Image | None = None
        self.width = 0
        self.height = 0

    def __enter__(self) -> "PillowRegionSource":
        self.image = Image.open(self.path)
        self.width, self.height = self.image.size
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        if self.image is not None:
            self.image.close()
            self.image = None

    def read_region(
        self, top: int, left: int, height: int, width: int
    ) -> np.ndarray:
        assert self.image is not None
        crop = self.image.crop((left, top, left + width, top + height)).convert(
            "RGB"
        )
        try:
            return np.array(crop, dtype=np.uint8, copy=True)
        finally:
            crop.close()


@contextmanager
def open_original_source(
    path: Path,
    scene_index: int | None,
) -> Iterator[object]:
    if path.suffix.lower() != ".czi":
        with PillowRegionSource(path) as source:
            yield source
        return
    try:
        from predict import open_image_sources
    except ImportError as error:
        raise RuntimeError(
            "CZI texture analysis requires the prediction dependencies from "
            "requirements.txt."
        ) from error
    with open_image_sources(path, scene_index) as sources:
        if scene_index is not None:
            matches = [source for index, source in sources if index == scene_index]
        elif len(sources) == 1:
            matches = [sources[0][1]]
        else:
            matches = []
        if len(matches) != 1:
            raise RuntimeError(
                f"Could not select one CZI scene from {path} for scene "
                f"{scene_index!r}"
            )
        yield matches[0]


def load_original_at_analysis_scale(
    scan: ScanInputs,
    native_height: int,
    native_width: int,
    analysis_height: int,
    analysis_width: int,
    chunk_rows: int,
) -> np.ndarray | None:
    if scan.original_path is None:
        return None
    started = time.perf_counter()
    log(f"{scan.name}: reading matching original {scan.original_path.name}")
    with open_original_source(scan.original_path, scan.original_scene) as source:
        if (source.height, source.width) != (native_height, native_width):
            raise ValueError(
                f"Original {scan.original_path} has dimensions "
                f"{source.width}x{source.height}; segmentation is "
                f"{native_width}x{native_height}"
            )
        output = np.empty(
            (analysis_height, analysis_width, 3), dtype=np.uint8
        )
        chunk_count = math.ceil(analysis_height / chunk_rows)
        for out_top in range(0, analysis_height, chunk_rows):
            out_bottom = min(out_top + chunk_rows, analysis_height)
            chunk_index = out_top // chunk_rows + 1
            if (
                chunk_index == 1
                or chunk_index == chunk_count
                or chunk_index % max(1, chunk_count // 10) == 0
            ):
                update_status(
                    f"{scan.name} | original RGB | chunk "
                    f"{chunk_index}/{chunk_count}"
                )
            native_top = math.floor(out_top * native_height / analysis_height)
            native_bottom = math.ceil(
                out_bottom * native_height / analysis_height
            )
            native_bottom = max(
                native_top + 1, min(native_bottom, native_height)
            )
            region = source.read_region(
                native_top,
                0,
                native_bottom - native_top,
                native_width,
            )
            region_image = Image.fromarray(region)
            try:
                resized = region_image.resize(
                    (analysis_width, out_bottom - out_top),
                    Image.Resampling.LANCZOS,
                    reducing_gap=2.0,
                )
                try:
                    output[out_top:out_bottom] = np.asarray(
                        resized, dtype=np.uint8
                    )
                finally:
                    resized.close()
            finally:
                region_image.close()
                del region
    log(
        "Original image loaded at analysis scale | "
        f"scan={scan.name} | source={scan.original_path} | "
        f"analysis_shape={output.shape} | bytes={output.nbytes:,} | "
        f"elapsed_seconds={time.perf_counter() - started:.3f}"
    )
    return output


def masked_glcm_metrics(
    gray: np.ndarray,
    mask: np.ndarray,
    levels: int,
) -> dict[str, float | None]:
    if np.count_nonzero(mask) < 2:
        return {
            "GLCM contrast": None,
            "GLCM homogeneity": None,
            "GLCM energy": None,
            "GLCM correlation": None,
        }
    quantized = np.minimum(
        (gray.astype(np.uint16) * levels) // 256, levels - 1
    ).astype(np.int32)
    matrix = np.zeros((levels, levels), dtype=np.float64)
    offsets = ((0, 1), (1, 0), (1, 1), (1, -1))
    for row_offset, col_offset in offsets:
        if col_offset >= 0:
            source_slice = (
                slice(0, gray.shape[0] - row_offset or None),
                slice(0, gray.shape[1] - col_offset or None),
            )
            target_slice = (
                slice(row_offset, None),
                slice(col_offset, None),
            )
        else:
            source_slice = (
                slice(0, gray.shape[0] - row_offset or None),
                slice(-col_offset, None),
            )
            target_slice = (
                slice(row_offset, None),
                slice(0, col_offset),
            )
        valid = mask[source_slice] & mask[target_slice]
        if not np.any(valid):
            continue
        first = quantized[source_slice][valid]
        second = quantized[target_slice][valid]
        counts = np.bincount(
            first * levels + second, minlength=levels * levels
        ).reshape((levels, levels))
        matrix += counts + counts.T
    total = matrix.sum()
    if total == 0:
        return {
            "GLCM contrast": None,
            "GLCM homogeneity": None,
            "GLCM energy": None,
            "GLCM correlation": None,
        }
    probability = matrix / total
    indices = np.arange(levels, dtype=np.float64)
    row_index, col_index = np.meshgrid(indices, indices, indexing="ij")
    contrast = np.sum(probability * (row_index - col_index) ** 2)
    homogeneity = np.sum(probability / (1.0 + (row_index - col_index) ** 2))
    energy = np.sum(probability ** 2)
    mean_row = np.sum(probability * row_index)
    mean_col = np.sum(probability * col_index)
    variance_row = np.sum(probability * (row_index - mean_row) ** 2)
    variance_col = np.sum(probability * (col_index - mean_col) ** 2)
    denominator = math.sqrt(variance_row * variance_col)
    correlation = (
        np.sum(
            probability
            * (row_index - mean_row)
            * (col_index - mean_col)
        ) / denominator
        if denominator > 0 else None
    )
    return {
        "GLCM contrast": float(contrast),
        "GLCM homogeneity": float(homogeneity),
        "GLCM energy": float(energy),
        "GLCM correlation": (
            float(correlation) if correlation is not None else None
        ),
    }


def cuda_dense_texture_features(
    scan_name: str,
    gray_float: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray] | None:
    """Calculate dense derivative features with tiled CuPy ndimage filters."""
    stage = "cupy_dense_texture"
    height, width = gray_float.shape
    if not ACCELERATOR.should_use_cupy(gray_float.size, stage):
        return None
    cupy = ACCELERATOR.cupy
    cupy_ndi = ACCELERATOR.cupy_ndi
    # The Hessian uses two sigma/sqrt(2) Gaussian-derivative passes with a
    # truncate of 100. Each pass has a 71-pixel radius at sigma=1, so 144 rows
    # of overlap safely preserve values at internal tile edges.
    halo_rows = 144
    chunk_rows = ACCELERATOR.choose_chunk_rows(
        width,
        height,
        bytes_per_pixel=96,
        halo_rows=halo_rows,
    )
    gradient_magnitude = np.empty_like(gray_float, dtype=np.float32)
    tensor_coherence = np.empty_like(gray_float, dtype=np.float32)
    tensor_orientation = np.empty_like(gray_float, dtype=np.float32)
    ridge_strength = np.empty_like(gray_float, dtype=np.float32)
    chunk_count = math.ceil(height / chunk_rows)
    log(
        f"{scan_name}: CuPy dense texture filters | chunk_rows={chunk_rows:,} | "
        f"halo_rows={halo_rows} | chunks={chunk_count:,}",
        logging.DEBUG,
    )
    try:
        for top in range(0, height, chunk_rows):
            bottom = min(top + chunk_rows, height)
            chunk_index = top // chunk_rows + 1
            read_top = max(0, top - halo_rows)
            read_bottom = min(height, bottom + halo_rows)
            local_top = top - read_top
            local_bottom = local_top + bottom - top
            update_status(
                f"{scan_name} | CUDA texture filters | "
                f"chunk {chunk_index}/{chunk_count}"
            )
            gpu_gray = cupy.asarray(
                np.ascontiguousarray(gray_float[read_top:read_bottom]),
                dtype=cupy.float32,
            )

            gradient_y = cupy_ndi.sobel(
                gpu_gray, axis=0, mode="reflect"
            ) / cupy.float32(8.0)
            gradient_x = cupy_ndi.sobel(
                gpu_gray, axis=1, mode="reflect"
            ) / cupy.float32(8.0)
            gpu_gradient_magnitude = cupy.hypot(
                gradient_y, gradient_x
            ).astype(cupy.float32, copy=False)

            derivative_r = cupy_ndi.sobel(
                gpu_gray, axis=0, mode="constant", cval=0
            )
            derivative_c = cupy_ndi.sobel(
                gpu_gray, axis=1, mode="constant", cval=0
            )
            tensor_rr = cupy_ndi.gaussian_filter(
                derivative_r * derivative_r,
                sigma=1.0,
                mode="constant",
                cval=0,
            )
            tensor_rc = cupy_ndi.gaussian_filter(
                derivative_r * derivative_c,
                sigma=1.0,
                mode="constant",
                cval=0,
            )
            tensor_cc = cupy_ndi.gaussian_filter(
                derivative_c * derivative_c,
                sigma=1.0,
                mode="constant",
                cval=0,
            )
            tensor_delta = cupy.sqrt(
                (tensor_rr - tensor_cc) ** 2
                + cupy.float32(4.0) * tensor_rc ** 2
            )
            gpu_tensor_coherence = tensor_delta / cupy.maximum(
                tensor_rr + tensor_cc, cupy.float32(EPSILON)
            )
            gpu_tensor_orientation = cupy.float32(0.5) * cupy.arctan2(
                cupy.float32(2.0) * tensor_rc,
                tensor_rr - tensor_cc,
            )

            sigma_scaled = (1.0 / math.sqrt(2.0),) * 2
            gaussian_kwargs = {
                "sigma": sigma_scaled,
                "mode": "constant",
                "cval": 0,
                "truncate": 100,
            }
            hessian_gradient_r = cupy_ndi.gaussian_filter(
                gpu_gray, order=(1, 0), **gaussian_kwargs
            )
            hessian_gradient_c = cupy_ndi.gaussian_filter(
                gpu_gray, order=(0, 1), **gaussian_kwargs
            )
            hessian_rr = cupy_ndi.gaussian_filter(
                hessian_gradient_r, order=(1, 0), **gaussian_kwargs
            )
            hessian_rc = cupy_ndi.gaussian_filter(
                hessian_gradient_r, order=(0, 1), **gaussian_kwargs
            )
            hessian_cc = cupy_ndi.gaussian_filter(
                hessian_gradient_c, order=(0, 1), **gaussian_kwargs
            )
            hessian_half_delta = cupy.sqrt(
                hessian_rc ** 2
                + ((hessian_rr - hessian_cc) / cupy.float32(2.0)) ** 2
            )
            hessian_minimum = (
                (hessian_rr + hessian_cc) / cupy.float32(2.0)
                - hessian_half_delta
            )
            gpu_ridge_strength = cupy.maximum(
                cupy.float32(0.0), -hessian_minimum
            ).astype(cupy.float32, copy=False)

            target_slice = slice(top, bottom)
            local_slice = slice(local_top, local_bottom)
            gradient_magnitude[target_slice] = cupy.asnumpy(
                gpu_gradient_magnitude[local_slice]
            )
            tensor_coherence[target_slice] = cupy.asnumpy(
                gpu_tensor_coherence[local_slice]
            )
            tensor_orientation[target_slice] = cupy.asnumpy(
                gpu_tensor_orientation[local_slice]
            )
            ridge_strength[target_slice] = cupy.asnumpy(
                gpu_ridge_strength[local_slice]
            )
            del (
                gpu_gray,
                gradient_y,
                gradient_x,
                gpu_gradient_magnitude,
                derivative_r,
                derivative_c,
                tensor_rr,
                tensor_rc,
                tensor_cc,
                tensor_delta,
                gpu_tensor_coherence,
                gpu_tensor_orientation,
                hessian_gradient_r,
                hessian_gradient_c,
                hessian_rr,
                hessian_rc,
                hessian_cc,
                hessian_half_delta,
                hessian_minimum,
                gpu_ridge_strength,
            )
        log(
            f"{scan_name}: CuPy dense texture filters complete | "
            f"chunks={chunk_count:,}",
            logging.DEBUG,
        )
        return (
            gradient_magnitude,
            tensor_coherence,
            tensor_orientation,
            ridge_strength,
        )
    except Exception as error:
        ACCELERATOR.disable_stage(stage, error)
        del (
            gradient_magnitude,
            tensor_coherence,
            tensor_orientation,
            ridge_strength,
        )
        return None


def texture_measurements(
    scan_name: str,
    labels: np.ndarray,
    rgb: np.ndarray | None,
    region_rows: list[dict[str, object]],
    args: argparse.Namespace,
) -> list[dict[str, object]]:
    if rgb is None:
        return [
            {
                "Scan": scan_name,
                "Class": class_name,
                "Status": "Unavailable: no matching original RGB image",
                **{field: None for field in TEXTURE_METRIC_FIELDS},
            }
            for class_name in CLASS_NAMES
        ]
    started = time.perf_counter()
    log(f"{scan_name}: calculating RGB, stain, gradient, and texture metrics")
    update_status(f"{scan_name} | texture | grayscale and gradients")
    gray = np.rint(
        0.2126 * rgb[..., 0]
        + 0.7152 * rgb[..., 1]
        + 0.0722 * rgb[..., 2]
    ).astype(np.uint8)
    gray_float = gray.astype(np.float32) / np.float32(255.0)
    update_status(f"{scan_name} | texture | local entropy")
    local_entropy = filters.rank.entropy(
        gray, morphology.disk(args.texture_entropy_radius)
    ).astype(np.float32)

    dense_features = cuda_dense_texture_features(scan_name, gray_float)
    if dense_features is not None:
        (
            gradient_magnitude,
            tensor_coherence,
            tensor_orientation,
            ridge_strength,
        ) = dense_features
        log(
            f"{scan_name}: gradient, structure-tensor, and Hessian arrays "
            "calculated with CuPy CUDA",
            logging.DEBUG,
        )
    else:
        update_status(f"{scan_name} | texture | CPU Sobel gradients")
        gradient_y = ndi.sobel(gray_float, axis=0, mode="reflect") / 8.0
        gradient_x = ndi.sobel(gray_float, axis=1, mode="reflect") / 8.0
        gradient_magnitude = np.hypot(
            gradient_y, gradient_x
        ).astype(np.float32)

        update_status(f"{scan_name} | texture | CPU structure tensor")
        tensor_rr, tensor_rc, tensor_cc = feature.structure_tensor(
            gray_float, sigma=1.0, order="rc"
        )
        tensor_delta = np.sqrt(
            (tensor_rr - tensor_cc) ** 2 + 4.0 * tensor_rc ** 2
        )
        tensor_coherence = tensor_delta / np.maximum(
            tensor_rr + tensor_cc, EPSILON
        )
        tensor_orientation = 0.5 * np.arctan2(
            2.0 * tensor_rc, tensor_rr - tensor_cc
        )
        del tensor_rr, tensor_rc, tensor_cc, tensor_delta

        update_status(f"{scan_name} | texture | CPU Hessian ridge response")
        hessian_elements = feature.hessian_matrix(
            gray_float,
            sigma=1.0,
            order="rc",
            use_gaussian_derivatives=True,
        )
        hessian_eigenvalues = feature.hessian_matrix_eigvals(hessian_elements)
        ridge_strength = np.maximum(
            0.0, -np.minimum(hessian_eigenvalues[0], hessian_eigenvalues[1])
        ).astype(np.float32)
        del hessian_elements, hessian_eigenvalues

    rows = []
    retained_regions: dict[str, list[dict[str, object]]] = defaultdict(list)
    for region_row in region_rows:
        if region_row.get("Scan") == scan_name:
            retained_regions[str(region_row["Class"])].append(region_row)
    pixel_height_um = args.pixel_height_um * args.analysis_downsample
    pixel_width_um = args.pixel_width_um * args.analysis_downsample
    for class_id, class_name in enumerate(CLASS_NAMES):
        update_status(f"{scan_name} | texture | {class_name}")
        mask = labels == class_id
        count = int(np.count_nonzero(mask))
        if not count:
            rows.append(
                {
                    "Scan": scan_name,
                    "Class": class_name,
                    "Status": "Class absent at analysis scale",
                    **{field: None for field in TEXTURE_METRIC_FIELDS},
                }
            )
            continue
        selected_rgb = rgb[mask].astype(np.float32) / np.float32(255.0)
        optical_density = -np.log(
            np.maximum((selected_rgb * 255.0 + 1.0) / 256.0, EPSILON)
        )
        hed = color.rgb2hed(selected_rgb.reshape((-1, 1, 3)))[:, 0, :]
        intensity_values = gray[mask]
        histogram = np.bincount(intensity_values, minlength=256).astype(np.float64)
        histogram /= histogram.sum()
        nonzero = histogram > 0
        intensity_entropy = float(
            -np.sum(histogram[nonzero] * np.log2(histogram[nonzero]))
        )
        orientation_mean, orientation_regularity = axial_orientation_statistics(
            tensor_orientation[mask], tensor_coherence[mask]
        )
        rgb_std = selected_rgb.std(axis=0)
        row: dict[str, object] = {
            "Scan": scan_name,
            "Class": class_name,
            "Status": "Measured",
            "Analysis pixels": count,
            "Red mean (0-1)": float(selected_rgb[:, 0].mean()),
            "Green mean (0-1)": float(selected_rgb[:, 1].mean()),
            "Blue mean (0-1)": float(selected_rgb[:, 2].mean()),
            "Red SD (0-1)": float(rgb_std[0]),
            "Green SD (0-1)": float(rgb_std[1]),
            "Blue SD (0-1)": float(rgb_std[2]),
            "Combined RGB variation": float(np.linalg.norm(rgb_std)),
            "Red optical density mean": float(optical_density[:, 0].mean()),
            "Green optical density mean": float(optical_density[:, 1].mean()),
            "Blue optical density mean": float(optical_density[:, 2].mean()),
            "Hematoxylin stain mean": float(hed[:, 0].mean()),
            "Eosin stain mean": float(hed[:, 1].mean()),
            "Residual DAB stain mean": float(hed[:, 2].mean()),
            "Grayscale mean (0-1)": float(gray_float[mask].mean()),
            "Grayscale SD (0-1)": float(gray_float[mask].std()),
            "Intensity Shannon entropy (bits)": intensity_entropy,
            "Gradient magnitude mean": float(gradient_magnitude[mask].mean()),
            "Gradient magnitude SD": float(gradient_magnitude[mask].std()),
            "Local entropy mean (bits)": float(local_entropy[mask].mean()),
            "Local entropy SD (bits)": float(local_entropy[mask].std()),
            "Structure-tensor coherence mean": float(
                tensor_coherence[mask].mean()
            ),
            "Structure-tensor orientation (degrees)": orientation_mean,
            "Striation-orientation regularity": orientation_regularity,
            "Hessian ridge strength mean": float(ridge_strength[mask].mean()),
            "Hessian ridge strength P90": float(
                np.percentile(ridge_strength[mask], 90)
            ),
            **masked_glcm_metrics(gray, mask, args.texture_levels),
        }
        rows.append(row)
        log(
            "Texture class complete | "
            f"scan={scan_name} | class={class_name} | pixels={count:,} | "
            f"rgb_variation={row['Combined RGB variation']:.6f} | "
            f"local_entropy_mean={row['Local entropy mean (bits)']:.6f} | "
            f"structure_coherence="
            f"{row['Structure-tensor coherence mean']:.6f}",
            logging.DEBUG,
        )

        retained = retained_regions.get(class_name, [])
        if retained:
            component_labels, _ = accelerated_label(
                mask, structure=np.ones((3, 3), dtype=np.uint8)
            )
            component_properties = measure.regionprops(component_labels)
            properties_by_centroid = {
                (
                    round(float(prop.centroid[0]), 8),
                    round(float(prop.centroid[1]), 8),
                ): prop
                for prop in component_properties
            }
            for region_row in retained:
                target_centroid = (
                    float(region_row["Centroid Y (µm)"]) / pixel_height_um,
                    float(region_row["Centroid X (µm)"]) / pixel_width_um,
                )
                prop = properties_by_centroid.get(
                    (
                        round(target_centroid[0], 8),
                        round(target_centroid[1], 8),
                    )
                )
                if prop is None:
                    # GPU connected-component numbering is not required to
                    # match SciPy numbering. Centroids provide a stable
                    # cross-backend identity for retained regions.
                    prop = min(
                        component_properties,
                        key=lambda candidate: (
                            (candidate.centroid[0] - target_centroid[0]) ** 2
                            + (candidate.centroid[1] - target_centroid[1]) ** 2
                        ),
                    )
                region_id = int(prop.label)
                region_slice = prop.slice
                local_mask = component_labels[region_slice] == region_id
                local_rgb = rgb[region_slice][local_mask].astype(np.float32) / 255.0
                local_gray = gray[region_slice][local_mask]
                local_histogram = np.bincount(
                    local_gray, minlength=256
                ).astype(np.float64)
                local_histogram /= local_histogram.sum()
                local_nonzero = local_histogram > 0
                local_coherence = tensor_coherence[region_slice][local_mask]
                local_orientation = tensor_orientation[region_slice][local_mask]
                local_orientation_mean, local_orientation_regularity = (
                    axial_orientation_statistics(
                        local_orientation, local_coherence
                    )
                )
                region_row.update(
                    {
                        "RGB variation within region": float(
                            np.linalg.norm(local_rgb.std(axis=0))
                        ),
                        "Intensity entropy within region (bits)": float(
                            -np.sum(
                                local_histogram[local_nonzero]
                                * np.log2(local_histogram[local_nonzero])
                            )
                        ),
                        "Gradient mean within region": float(
                            gradient_magnitude[region_slice][local_mask].mean()
                        ),
                        "Local entropy mean within region (bits)": float(
                            local_entropy[region_slice][local_mask].mean()
                        ),
                        "Structure coherence within region": float(
                            local_coherence.mean()
                        ),
                        "Texture orientation within region (degrees)": (
                            local_orientation_mean
                        ),
                        "Texture-orientation regularity within region": (
                            local_orientation_regularity
                        ),
                        "Hessian ridge mean within region": float(
                            ridge_strength[region_slice][local_mask].mean()
                        ),
                    }
                )
            del component_labels, component_properties, properties_by_centroid
        del selected_rgb, optical_density, hed
    log(
        f"{scan_name}: texture measurements complete | "
        f"rows={len(rows)} | elapsed_seconds={time.perf_counter() - started:.3f}"
    )
    return rows


def load_boundary_probability(
    path: Path | None,
    analysis_height: int,
    analysis_width: int,
) -> np.ndarray | None:
    if path is None:
        return None
    with Image.open(path) as image:
        resized = image.resize(
            (analysis_width, analysis_height), Image.Resampling.BILINEAR
        ).convert("L")
        try:
            return (
                np.asarray(resized, dtype=np.float32)
                / np.float32(255.0)
            )
        finally:
            resized.close()


def analyze_scan(
    scan: ScanInputs,
    scan_index: int,
    scan_count: int,
    args: argparse.Namespace,
) -> AnalysisResults:
    scan_started = time.perf_counter()
    log(f"Scan {scan_index}/{scan_count}: {scan.name}")
    for input_type, path in (
        ("label", scan.label_path),
        ("probability", scan.probability_path),
        ("boundary", scan.boundary_path),
        ("original", scan.original_path),
    ):
        log(
            "Scan input | "
            f"scan={scan.name} | type={input_type} | "
            f"path={path or 'unavailable'} | "
            f"bytes={file_size_or_unavailable(path)}",
            logging.DEBUG,
        )
    for class_name, path in zip(CLASS_NAMES, scan.grayscale_paths):
        log(
            "Scan input | "
            f"scan={scan.name} | type=grayscale_probability | class={class_name} | "
            f"path={path or 'unavailable'} | "
            f"bytes={file_size_or_unavailable(path)}",
            logging.DEBUG,
        )
    log_resource_snapshot(f"scan_{scan_index}_start", args.output)
    labels, hard_counts, native_height, native_width = load_label_data(
        scan.label_path, args.analysis_downsample, args.chunk_rows
    )
    analysis_height, analysis_width = labels.shape
    log(
        f"{scan.name}: native={native_width}x{native_height}, "
        f"analysis={analysis_width}x{analysis_height}"
    )
    qc_rows: list[dict[str, object]] = []
    missing_classes = [
        class_name
        for class_name, count in zip(CLASS_NAMES, hard_counts)
        if count == 0
    ]
    if missing_classes:
        log(
            f"{scan.name}: missing hard-prediction classes: "
            + ", ".join(missing_classes),
            logging.WARNING,
        )
        qc_rows.append(
            {
                "Scan": scan.name,
                "Severity": "Warning",
                "Check": "Missing hard-prediction classes",
                "Result": ", ".join(missing_classes),
            }
        )
    if int(hard_counts[:-1].sum()) == 0:
        log(
            f"{scan.name}: all native-resolution pixels are Background",
            logging.ERROR,
        )
        qc_rows.append(
            {
                "Scan": scan.name,
                "Severity": "Critical",
                "Check": "All-background segmentation",
                "Result": "Every native-resolution pixel is Background",
            }
        )

    with ProbabilitySource(
        scan, native_height, native_width
    ) as probability_source:
        (
            probability_scan_metrics,
            probability_class_rows,
            uncertain_mask,
            bone_fibro_mask,
        ) = probability_statistics(
            probability_source,
            scan.label_path,
            labels,
            native_height,
            native_width,
            args,
        )
        if not probability_source.available:
            qc_rows.append(
                {
                    "Scan": scan.name,
                    "Severity": "Warning",
                    "Check": "Probability-derived measurements",
                    "Result": (
                        "Unavailable: neither a probability .npy nor all six "
                        "grayscale probability maps were found"
                    ),
                }
            )
        (
            class_rows,
            region_rows,
            pore_rows,
            centroids_by_class,
            interface_curvature,
            region_rows_omitted,
            pore_rows_omitted,
        ) = analyze_classes(
            scan,
            labels,
            hard_counts,
            probability_source,
            probability_class_rows,
            args,
        )

    analysis_pixel_height_um = args.pixel_height_um * args.analysis_downsample
    analysis_pixel_width_um = args.pixel_width_um * args.analysis_downsample
    stage_started = time.perf_counter()
    update_status(f"{scan.name} | class-pair interfaces")
    interface_rows, contact_lengths = pair_contact_measurements(
        scan.name,
        labels,
        interface_curvature,
        uncertain_mask,
        bone_fibro_mask,
        analysis_pixel_height_um,
        analysis_pixel_width_um,
    )
    log(
        f"{scan.name}: interface measurements complete | "
        f"rows={len(interface_rows)} | "
        f"elapsed_seconds={time.perf_counter() - stage_started:.3f}"
    )
    stage_started = time.perf_counter()
    update_status(f"{scan.name} | directed spatial distances")
    spatial_rows = spatial_measurements(
        scan.name,
        labels,
        centroids_by_class,
        contact_lengths,
        analysis_pixel_height_um,
        analysis_pixel_width_um,
        args.proximity_um,
    )
    log(
        f"{scan.name}: spatial-distance measurements complete | "
        f"rows={len(spatial_rows)} | "
        f"elapsed_seconds={time.perf_counter() - stage_started:.3f}"
    )
    stage_started = time.perf_counter()
    update_status(f"{scan.name} | distance bands from Bone")
    distance_band_rows = distance_bands_from_bone(
        scan.name,
        labels,
        analysis_pixel_height_um,
        analysis_pixel_width_um,
        args.distance_bin_um,
    )
    log(
        f"{scan.name}: Bone-distance bands complete | "
        f"rows={len(distance_band_rows)} | "
        f"elapsed_seconds={time.perf_counter() - stage_started:.3f}"
    )

    contact_total_by_class = defaultdict(float)
    adjacent_count_by_class = defaultdict(int)
    for row in interface_rows:
        first = CLASS_NAMES.index(str(row["Class A"]))
        second = CLASS_NAMES.index(str(row["Class B"]))
        length_um = float(row["Interface length (µm)"])
        contact_total_by_class[first] += length_um
        contact_total_by_class[second] += length_um
        if row["Adjacent"]:
            adjacent_count_by_class[first] += 1
            adjacent_count_by_class[second] += 1
    for class_id, row in enumerate(class_rows):
        row["Interclass contact length (µm)"] = contact_total_by_class[class_id]
        row["Adjacent class count"] = adjacent_count_by_class[class_id]

    update_status(f"{scan.name} | boundary confidence and uncertainty")
    hard_boundary = hard_boundary_mask(labels)
    uncertain_boundary_fraction = (
        float(np.count_nonzero(hard_boundary & uncertain_mask))
        / np.count_nonzero(hard_boundary)
        if uncertain_mask is not None and np.any(hard_boundary)
        else None
    )
    boundary_probability = load_boundary_probability(
        scan.boundary_path, analysis_height, analysis_width
    )
    boundary_model_mean = (
        float(boundary_probability.mean())
        if boundary_probability is not None else None
    )
    boundary_model_on_hard = (
        float(boundary_probability[hard_boundary].mean())
        if boundary_probability is not None and np.any(hard_boundary)
        else None
    )

    rgb = None
    original_error = None
    if scan.original_path is not None:
        try:
            rgb = load_original_at_analysis_scale(
                scan,
                native_height,
                native_width,
                analysis_height,
                analysis_width,
                args.chunk_rows,
            )
        except Exception as error:
            original_error = f"{type(error).__name__}: {error}"
            qc_rows.append(
                {
                    "Scan": scan.name,
                    "Severity": "Warning",
                    "Check": "Original-image texture measurements",
                    "Result": original_error,
                }
            )
            log(
                f"{scan.name}: original-image analysis skipped: {original_error}",
                logging.WARNING,
            )
    texture_rows = texture_measurements(
        scan.name, labels, rgb, region_rows, args
    )

    pixel_area_um2 = args.pixel_width_um * args.pixel_height_um
    total_pixels = int(hard_counts.sum())
    non_background_pixels = int(hard_counts[:-1].sum())
    scan_row: dict[str, object] = {
        "Scan": scan.name,
        "Native width (pixels)": native_width,
        "Native height (pixels)": native_height,
        "Analysis width (pixels)": analysis_width,
        "Analysis height (pixels)": analysis_height,
        "Analysis downsample": args.analysis_downsample,
        "Pixel width (µm)": args.pixel_width_um,
        "Pixel height (µm)": args.pixel_height_um,
        "Total area (mm²)": total_pixels * pixel_area_um2 / 1_000_000.0,
        "Hard tissue area (mm²)": (
            non_background_pixels * pixel_area_um2 / 1_000_000.0
        ),
        "Hard background fraction": (
            float(hard_counts[-1] / total_pixels) if total_pixels else None
        ),
        "Missing class count": len(missing_classes),
        "Missing classes": ", ".join(missing_classes),
        "All background": non_background_pixels == 0,
        "Boundary-map source": (
            str(scan.boundary_path) if scan.boundary_path else "Unavailable"
        ),
        "Boundary probability mean": boundary_model_mean,
        "Boundary probability on hard boundaries mean": boundary_model_on_hard,
        "Uncertain hard-boundary fraction": uncertain_boundary_fraction,
        "Uncertain hard-boundary area (µm²)": (
            float(np.count_nonzero(hard_boundary & uncertain_mask))
            * analysis_pixel_height_um
            * analysis_pixel_width_um
            if uncertain_mask is not None else None
        ),
        "Original image": (
            str(scan.original_path) if scan.original_path else "Unavailable"
        ),
        "Original texture status": (
            "Measured" if rgb is not None
            else (original_error or "Unavailable")
        ),
        **probability_scan_metrics,
    }
    qc_rows.append(
        {
            "Scan": scan.name,
            "Severity": "Information",
            "Check": "Overlap-tile variation",
            "Result": (
                "Unavailable from saved outputs because per-tile predictions "
                "are discarded after blending"
            ),
        }
    )
    qc_rows.append(
        {
            "Scan": scan.name,
            "Severity": "Information",
            "Check": "Porosity interpretation",
            "Result": (
                "Closed holes are enclosed non-class islands in each hard mask; "
                "they are not automatically equivalent to biological pores"
            ),
        }
    )
    del labels, hard_boundary, boundary_probability, rgb
    gc.collect()
    results = AnalysisResults(
        scan_rows=[scan_row],
        class_rows=class_rows,
        region_rows=region_rows,
        pore_rows=pore_rows,
        interface_rows=interface_rows,
        spatial_rows=spatial_rows,
        distance_band_rows=distance_band_rows,
        texture_rows=texture_rows,
        qc_rows=qc_rows,
        region_rows_omitted=region_rows_omitted,
        pore_rows_omitted=pore_rows_omitted,
    )
    log_resource_snapshot(f"scan_{scan_index}_complete", args.output)
    log(
        "Scan analysis complete | "
        f"scan={scan.name} | class_rows={len(class_rows)} | "
        f"region_rows={len(region_rows):,} | pore_rows={len(pore_rows):,} | "
        f"interface_rows={len(interface_rows)} | spatial_rows={len(spatial_rows)} | "
        f"distance_band_rows={len(distance_band_rows):,} | "
        f"texture_rows={len(texture_rows)} | qc_rows={len(qc_rows)} | "
        f"elapsed_seconds={time.perf_counter() - scan_started:.3f}"
    )
    return results


def extend_results(target: AnalysisResults, source: AnalysisResults) -> None:
    for attribute in (
        "scan_rows",
        "class_rows",
        "region_rows",
        "pore_rows",
        "interface_rows",
        "spatial_rows",
        "distance_band_rows",
        "texture_rows",
        "qc_rows",
    ):
        getattr(target, attribute).extend(getattr(source, attribute))
    target.region_rows_omitted += source.region_rows_omitted
    target.pore_rows_omitted += source.pore_rows_omitted


def metric_definitions(args: argparse.Namespace) -> list[dict[str, object]]:
    return [
        {
            "Metric or group": "Native-resolution measurements",
            "Definition": (
                "Hard pixel counts, hard areas, expected/probability-weighted "
                "areas, confidence, entropy, and top-two margins are calculated "
                "at the saved output's native dimensions."
            ),
            "Interpretation / limitation": "Not affected by --analysis-downsample.",
        },
        {
            "Metric or group": "Analysis-scale measurements",
            "Definition": (
                "Morphology, topology, interfaces, distance transforms, skeletons, "
                "and textures are calculated after nearest-neighbor label "
                f"downsampling by {args.analysis_downsample}."
            ),
            "Interpretation / limitation": (
                "Physical units are rescaled, but small regions and fine boundaries "
                "can be lost when downsample > 1."
            ),
        },
        {
            "Metric or group": "Hard reference-area fraction",
            "Definition": (
                "For tissue classes, class hard pixels divided by all non-Background "
                "hard pixels. For Background, Background pixels divided by all pixels."
            ),
            "Interpretation / limitation": "Exclusive hard segmentation.",
        },
        {
            "Metric or group": "Expected area",
            "Definition": "Sum of class probabilities multiplied by calibrated pixel area.",
            "Interpretation / limitation": (
                "Uses float32 Probabilities/*.npy when present; otherwise 8-bit "
                "Grayscale maps."
            ),
        },
        {
            "Metric or group": "Normalized entropy",
            "Definition": "Per-pixel Shannon entropy divided by log(6), then averaged.",
            "Interpretation / limitation": "0 is decisive; 1 is maximally uncertain.",
        },
        {
            "Metric or group": "Top-two margin",
            "Definition": (
                "Difference between the two largest normalized class probabilities. "
                f"Low margin means ≤ {args.uncertainty_margin:g}."
            ),
            "Interpretation / limitation": "Smaller values indicate ambiguity.",
        },
        {
            "Metric or group": "Bone-Fibrocartilage ambiguity",
            "Definition": (
                "Pixels whose two most likely classes are Bone and Fibrocartilage "
                "and whose top-two margin is below the configured threshold."
            ),
            "Interpretation / limitation": "Probability-based transition-zone proxy.",
        },
        {
            "Metric or group": "Connected region",
            "Definition": "An 8-connected component of one hard class mask.",
            "Interpretation / limitation": "One biological object may be fragmented by errors.",
        },
        {
            "Metric or group": "Circularity",
            "Definition": "4π × area / perimeter².",
            "Interpretation / limitation": "Near 1 for compact round regions; resolution-sensitive.",
        },
        {
            "Metric or group": "Solidity",
            "Definition": "Region area divided by convex-hull area.",
            "Interpretation / limitation": "Low values indicate indentations or fragmentation.",
        },
        {
            "Metric or group": "Convexity",
            "Definition": "Convex-hull perimeter divided by region perimeter.",
            "Interpretation / limitation": "Near 1 for smooth convex regions.",
        },
        {
            "Metric or group": "Boundary roughness",
            "Definition": "Region perimeter divided by convex-hull perimeter.",
            "Interpretation / limitation": "Reciprocal-style companion to convexity.",
        },
        {
            "Metric or group": "Fragmentation fraction",
            "Definition": "1 minus largest connected-region area / total class area.",
            "Interpretation / limitation": "0 means all class area is one component.",
        },
        {
            "Metric or group": "Clark-Evans clustering index",
            "Definition": (
                "Observed mean same-class centroid nearest-neighbor distance divided "
                "by the Poisson expectation 0.5/sqrt(region density)."
            ),
            "Interpretation / limitation": "<1 clustered; >1 dispersed; edge effects are uncorrected.",
        },
        {
            "Metric or group": "Closed hole / pore",
            "Definition": (
                "An 8-connected non-class island enclosed after binary hole filling "
                "of one class mask."
            ),
            "Interpretation / limitation": (
                "May be another tissue class or segmentation error; not automatically "
                "a biological pore."
            ),
        },
        {
            "Metric or group": "Euler number",
            "Definition": "Number of connected regions minus number of closed holes.",
            "Interpretation / limitation": "Topology is highly sensitive to small mask errors.",
        },
        {
            "Metric or group": "Skeleton metrics",
            "Definition": (
                "8-connected medial skeleton length with endpoint pixels and clustered "
                "junction pixels."
            ),
            "Interpretation / limitation": "Branches can multiply in noisy boundaries.",
        },
        {
            "Metric or group": "Local thickness",
            "Definition": "Twice the physical Euclidean distance to the mask edge at skeleton pixels.",
            "Interpretation / limitation": "A medial-axis thickness estimate.",
        },
        {
            "Metric or group": "Internal-void skeleton density",
            "Definition": "Skeleton length of enclosed holes divided by hard tissue area.",
            "Interpretation / limitation": "A fracture/crack proxy, not a validated crack classifier.",
        },
        {
            "Metric or group": "Interface length",
            "Definition": "Sum of physical lengths of horizontal/vertical pixel edges shared by two classes.",
            "Interpretation / limitation": "4-neighbor edge estimator at analysis scale.",
        },
        {
            "Metric or group": "Interface roughness proxy",
            "Definition": "Interface length divided by the diagonal of its overall bounding box.",
            "Interpretation / limitation": "Can increase with multiple disjoint interfaces.",
        },
        {
            "Metric or group": "Interface curvature",
            "Definition": "Mean absolute divergence of the signed-distance unit normal on both class sides.",
            "Interpretation / limitation": "Skipped with --skip-curvature; resolution-sensitive.",
        },
        {
            "Metric or group": "Transition-width proxy",
            "Definition": "Bone-Fibrocartilage ambiguous area divided by their hard interface length.",
            "Interpretation / limitation": "Undefined when the hard masks do not touch.",
        },
        {
            "Metric or group": "Pixel-to-target distance",
            "Definition": "Physical Euclidean distance from every source-class pixel center to target-class pixels.",
            "Interpretation / limitation": "Directed; source→target need not equal target→source.",
        },
        {
            "Metric or group": "Distance bands from Bone",
            "Definition": f"Class area binned every {args.distance_bin_um:g} µm from the Bone mask.",
            "Interpretation / limitation": "Uses analysis-scale labels.",
        },
        {
            "Metric or group": "RGB and optical density",
            "Definition": "Class-masked original RGB means/SDs and -log((channel+1)/256).",
            "Interpretation / limitation": "Requires a dimension-matched original raster or CZI scene.",
        },
        {
            "Metric or group": "H/E stain values",
            "Definition": "scikit-image RGB-to-HED color deconvolution, reported as class means.",
            "Interpretation / limitation": "Stain-vector assumptions should be validated for the acquisition protocol.",
        },
        {
            "Metric or group": "GLCM texture",
            "Definition": (
                "Masked, symmetric co-occurrence matrix at one-pixel offsets in "
                "0°, 45°, 90°, and 135° directions."
            ),
            "Interpretation / limitation": f"Uses {args.texture_levels} gray levels at analysis scale.",
        },
        {
            "Metric or group": "Structure tensor",
            "Definition": "Gaussian-smoothed local gradient tensor with mean coherence and axial orientation.",
            "Interpretation / limitation": "Orientation regularity near 1 means aligned texture.",
        },
        {
            "Metric or group": "Hessian ridge strength",
            "Definition": "Positive response from the negative minimum Hessian eigenvalue at sigma=1.",
            "Interpretation / limitation": "Scale-dependent ridge/striation proxy.",
        },
        {
            "Metric or group": "Overlap-tile variation",
            "Definition": "Not calculated.",
            "Interpretation / limitation": (
                "Per-tile predictions are discarded after blending and cannot be "
                "recovered from the saved outputs."
            ),
        },
    ]


def ordered_headers(records: list[dict[str, object]]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for record in records:
        for key in record:
            if key not in seen:
                headers.append(key)
                seen.add(key)
    return headers


def excel_number_format(header: str) -> str:
    lower = header.casefold()
    if "fraction" in lower:
        return "0.00%"
    if "area (mm²)" in lower:
        return "0.000000"
    if "area (µm²)" in lower:
        return "#,##0.00"
    if "(µm)" in lower or "(1/µm)" in lower:
        return "#,##0.000"
    if any(
        token in lower
        for token in ("pixels", "count", "region id", "pore id", "class id")
    ):
        return "#,##0"
    if any(
        token in lower
        for token in (
            "confidence", "entropy", "circularity", "solidity", "convexity",
            "roughness", "eccentricity", "aspect ratio", "elongation", "extent",
            "regularity", "coherence", "correlation", "energy", "homogeneity",
            "contrast", "index", "concentration", "mean (0-1)", "sd (0-1)",
            "optical density", "stain mean", "gradient", "ridge strength",
        )
    ):
        return "0.0000"
    return "0.000000"


def write_records_sheet(
    workbook: Workbook,
    name: str,
    records: list[dict[str, object]],
    table_index: int,
) -> None:
    started = time.perf_counter()
    update_status(f"Workbook | {name} | {len(records):,} row(s)")
    log(
        f"Writing worksheet | name={name} | records={len(records):,}",
        logging.DEBUG,
    )
    worksheet = workbook.create_sheet(name)
    worksheet.sheet_view.showGridLines = False
    if not records:
        worksheet.append(["Status"])
        worksheet.append(["No records were generated"])
        worksheet.freeze_panes = "A2"
        return
    headers = ordered_headers(records)
    worksheet.append(headers)
    for record in records:
        worksheet.append([finite_or_none(record.get(header)) for header in headers])

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)
    worksheet.row_dimensions[1].height = 32
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    table_name = re.sub(r"[^A-Za-z0-9_]", "", f"Table{table_index}{name}")[:250]
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    class_columns = [
        index
        for index, header in enumerate(headers, start=1)
        if header in {"Class", "Class A", "Class B", "Source class", "Target class"}
    ]
    for row in range(2, worksheet.max_row + 1):
        for column in class_columns:
            cell = worksheet.cell(row=row, column=column)
            class_name = str(cell.value)
            if class_name in CLASS_COLORS:
                cell.fill = PatternFill("solid", fgColor=CLASS_COLORS[class_name])
                if class_name in {"Bone", "Cartilage", "Marrow"}:
                    cell.font = Font(color="FFFFFF")

    for column_index, header in enumerate(headers, start=1):
        number_format = excel_number_format(header)
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column_index)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right")
        sample_values = [header]
        for row in range(2, min(worksheet.max_row, 200) + 1):
            value = worksheet.cell(row=row, column=column_index).value
            sample_values.append("" if value is None else str(value))
        maximum = max(len(value) for value in sample_values)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(maximum + 2, 11), 42
        )
    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 18

    fraction_columns = [
        index
        for index, header in enumerate(headers, start=1)
        if "fraction" in header.casefold()
    ]
    for column in fraction_columns:
        letter = get_column_letter(column)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{worksheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F7FBFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="9ECAE1",
                end_type="max",
                end_color="08519C",
            ),
        )

    freeze_locations = {
        "Scan Summary": "B2",
        "Class Summary": "D2",
        "Region Details": "D2",
        "Pore Details": "D2",
        "Interfaces": "D2",
        "Spatial Distances": "D2",
        "Distance Bands": "C2",
        "Texture Color": "C2",
        "QC": "C2",
        "Metric Definitions": "B2",
    }
    worksheet.freeze_panes = freeze_locations.get(name, "A2")
    if name == "QC":
        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 16
        worksheet.column_dimensions["C"].width = 34
        worksheet.column_dimensions["D"].width = 100
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 36
            for column in range(1, 5):
                worksheet.cell(row, column).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
    elif name == "Metric Definitions":
        worksheet.column_dimensions["A"].width = 34
        worksheet.column_dimensions["B"].width = 80
        worksheet.column_dimensions["C"].width = 70
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 45
            for column in range(1, 4):
                worksheet.cell(row, column).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
    log(
        f"Worksheet complete | name={name} | rows={worksheet.max_row:,} | "
        f"columns={worksheet.max_column:,} | "
        f"elapsed_seconds={time.perf_counter() - started:.3f}",
        logging.DEBUG,
    )


def write_readme_sheet(
    workbook: Workbook,
    args: argparse.Namespace,
    scans: list[ScanInputs],
    results: AnalysisResults,
) -> None:
    worksheet = workbook.active
    worksheet.title = "README"
    worksheet.sheet_view.showGridLines = False
    rows = [
        ("Report", "Tissue Segmentation Quantification 3"),
        ("Generated", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("Prediction directory", str(args.prediction_dir.resolve())),
        ("Scan count", len(scans)),
        ("Pixel width (µm)", args.pixel_width_um),
        ("Pixel height (µm)", args.pixel_height_um),
        ("Analysis downsample", args.analysis_downsample),
        ("Requested compute device", args.device),
        ("Resolved compute device", args.resolved_device),
        ("CUDA device", args.cuda_device_name),
        ("PyTorch CUDA enabled", args.pytorch_cuda_enabled),
        ("CuPy ndimage enabled", args.cupy_enabled),
        ("GPU memory fraction", args.gpu_memory_fraction),
        ("GPU chunk rows", args.gpu_chunk_rows or "Automatic"),
        ("GPU minimum pixels", args.gpu_min_pixels),
        ("GPU fallback count", ACCELERATOR.fallback_count),
        ("Low-confidence threshold", args.low_confidence_threshold),
        ("Uncertainty margin", args.uncertainty_margin),
        ("High normalized-entropy threshold", args.high_entropy_threshold),
        ("Proximity distance (µm)", args.proximity_um),
        ("Distance-band width (µm)", args.distance_bin_um),
        ("Region detail rows omitted", results.region_rows_omitted),
        ("Pore detail rows omitted", results.pore_rows_omitted),
        ("Detailed notes", "Quantification 3 Notes.md beside the script"),
        (
            "Important",
            "Validate calibration and segmentation accuracy before biological interpretation.",
        ),
    ]
    worksheet.append(["Setting", "Value"])
    for key, value in rows:
        worksheet.append([key, finite_or_none(value)])
    worksheet.append([])
    worksheet.append(["Sheet", "Contents"])
    sheet_descriptions = [
        ("Scan Summary", "One row per scan with calibration, area, uncertainty, and QC overview."),
        ("Class Summary", "One row per scan/class with area, morphology, topology, skeleton, and confidence metrics."),
        ("Region Details", "Largest retained connected regions; summaries still use all regions."),
        ("Pore Details", "Largest retained enclosed holes; summaries still use all holes."),
        ("Interfaces", "Unordered class-pair contact, curvature, roughness, and uncertainty."),
        ("Spatial Distances", "Directed source-to-target pixel and centroid distances."),
        ("Distance Bands", "Class-area distribution in distance bands from Bone."),
        ("Texture Color", "Optional original-image color, stain, gradient, and texture measurements."),
        ("QC", "Warnings, unavailable measurements, and interpretation cautions."),
        ("Metric Definitions", "Definitions and limitations for major metric groups."),
    ]
    for item in sheet_descriptions:
        worksheet.append(item)
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    second_header_row = len(rows) + 3
    for cell in worksheet[second_header_row]:
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.font = Font(color="FFFFFF", bold=True)
    worksheet.column_dimensions["A"].width = 36
    worksheet.column_dimensions["B"].width = 100
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"


def build_workbook(
    args: argparse.Namespace,
    scans: list[ScanInputs],
    results: AnalysisResults,
) -> None:
    started = time.perf_counter()
    log(
        "Workbook build started | "
        f"output={args.output.resolve()} | scans={len(scans):,} | "
        f"scan_rows={len(results.scan_rows):,} | "
        f"class_rows={len(results.class_rows):,} | "
        f"region_rows={len(results.region_rows):,} | "
        f"pore_rows={len(results.pore_rows):,} | "
        f"interface_rows={len(results.interface_rows):,} | "
        f"spatial_rows={len(results.spatial_rows):,} | "
        f"distance_band_rows={len(results.distance_band_rows):,} | "
        f"texture_rows={len(results.texture_rows):,} | qc_rows={len(results.qc_rows):,}"
    )
    workbook = Workbook()
    write_readme_sheet(workbook, args, scans, results)
    sheets = [
        ("Scan Summary", results.scan_rows),
        ("Class Summary", results.class_rows),
        ("Region Details", results.region_rows),
        ("Pore Details", results.pore_rows),
        ("Interfaces", results.interface_rows),
        ("Spatial Distances", results.spatial_rows),
        ("Distance Bands", results.distance_band_rows),
        ("Texture Color", results.texture_rows),
        ("QC", results.qc_rows),
        ("Metric Definitions", metric_definitions(args)),
    ]
    for table_index, (name, records) in enumerate(sheets, start=1):
        if len(records) + 1 > EXCEL_MAX_ROWS:
            raise RuntimeError(
                f"{name} would contain {len(records):,} data rows, exceeding "
                "Excel's worksheet limit. Reduce the detail-row flags."
            )
        write_records_sheet(workbook, name, records, table_index)
    for worksheet in workbook.worksheets:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.tabColor = "5B9BD5"
    args.output.parent.mkdir(parents=True, exist_ok=True)
    update_status(f"Workbook | saving {args.output.name}")
    workbook.save(args.output)
    workbook.close()
    log(
        "Workbook saved | "
        f"path={args.output.resolve()} | bytes={file_size_or_unavailable(args.output)} | "
        f"elapsed_seconds={time.perf_counter() - started:.3f}"
    )


def verify_workbook(path: Path, expected_scans: int) -> None:
    started = time.perf_counter()
    update_status(f"Workbook | validating {path.name}")
    log(f"Workbook validation started | path={path.resolve()}", logging.DEBUG)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        required = {
            "README", "Scan Summary", "Class Summary", "Region Details",
            "Pore Details", "Interfaces", "Spatial Distances", "Distance Bands",
            "Texture Color", "QC", "Metric Definitions",
        }
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise RuntimeError(f"Workbook is missing sheets: {sorted(missing)}")
        scan_sheet = workbook["Scan Summary"]
        if scan_sheet.max_row != expected_scans + 1:
            raise RuntimeError(
                f"Scan Summary has {scan_sheet.max_row - 1} rows; expected "
                f"{expected_scans}"
            )
        class_sheet = workbook["Class Summary"]
        if class_sheet.max_row != expected_scans * len(CLASS_NAMES) + 1:
            raise RuntimeError("Class Summary row count does not match scans × classes")
        headers = [cell.value for cell in class_sheet[1]]
        area_column = headers.index("Hard area (mm²)") + 1
        value = class_sheet.cell(row=2, column=area_column).value
        if not isinstance(value, (int, float)):
            raise RuntimeError("Numeric workbook values were written as text")
        log(
            "Workbook validation passed | "
            f"sheets={len(workbook.sheetnames)} | scan_rows={scan_sheet.max_row - 1} | "
            f"class_rows={class_sheet.max_row - 1} | numeric_sample={value} | "
            f"elapsed_seconds={time.perf_counter() - started:.3f}"
        )
    finally:
        workbook.close()


def main() -> None:
    args = parse_args()
    configure_logging(args)
    process_started = time.perf_counter()
    try:
        ACCELERATOR.configure(args)
        log_startup(args)
        log("Step 1/5: discovering scans and optional inputs")
        scans = discover_scans(args.prediction_dir, args.original_dir)
        log(f"Discovered {len(scans)} scan(s)")
        combined = AnalysisResults([], [], [], [], [], [], [], [], [])
        region_limit = args.max_region_rows
        pore_limit = args.max_pore_rows
        log("Step 2/5: calculating measurements")
        for index, scan in enumerate(scans, start=1):
            scan_args = argparse.Namespace(**vars(args))
            scan_args.max_region_rows = max(
                0, region_limit - len(combined.region_rows)
            )
            scan_args.max_pore_rows = max(
                0, pore_limit - len(combined.pore_rows)
            )
            scan_results = analyze_scan(scan, index, len(scans), scan_args)
            extend_results(combined, scan_results)
            ACCELERATOR.release_memory()
        log("Step 3/5: recording every measurement in the detailed log")
        log_all_measurements(combined, args.measurement_log_mode)
        log_resource_snapshot("before_workbook", args.output)
        log("Step 4/5: writing formatted Excel workbook")
        build_workbook(args, scans, combined)
        log("Step 5/5: validating workbook structure and numeric cell types")
        verify_workbook(args.output, len(scans))
        log_resource_snapshot("process_complete", args.output)
        log(
            f"Saved and verified: {args.output.resolve()} | "
            f"elapsed_seconds={time.perf_counter() - process_started:.3f} | "
            f"gpu_fallbacks={ACCELERATOR.fallback_count} | "
            f"log={args.log_file.resolve()}"
        )
    except Exception:
        STATUS_LINE.clear()
        ACCELERATOR.release_memory()
        log_resource_snapshot("process_failure", args.output)
        LOGGER.exception(
            "Quantification failed after %.3f seconds; detailed traceback follows",
            time.perf_counter() - process_started,
        )
        log(f"Failure log retained at: {args.log_file.resolve()}")
        raise


if __name__ == "__main__":
    main()
